import sqlite3
import sys, os, io, json, secrets
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, send_from_directory, abort
from werkzeug.security import generate_password_hash, check_password_hash
import barcode
from barcode.writer import SVGWriter
try:
    from dotenv import load_dotenv
except Exception:
    load_dotenv = None

app = Flask(__name__)
BASE = os.path.dirname(os.path.abspath(sys.executable if getattr(sys, 'frozen', False) else __file__))
if load_dotenv:
    try:
        load_dotenv(os.path.join(BASE, '.env'))
    except Exception:
        pass
DB = os.path.join(BASE, 'inventory.db')
EXPORTS_DIR = os.path.join(BASE, 'exports')
os.makedirs(EXPORTS_DIR, exist_ok=True)

def get_secret_key():
    env_key = os.environ.get('SECRET_KEY')
    if env_key:
        return env_key
    key_file = os.path.join(BASE, 'secret_key.txt')
    try:
        with open(key_file) as f:
            return f.read().strip()
    except Exception:
        pass
    key = os.urandom(24).hex()
    try:
        with open(key_file, 'w') as f:
            f.write(key)
    except Exception:
        pass
    return key

app.secret_key = get_secret_key()

def ensure_csrf_token():
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_urlsafe(32)
    return session['csrf_token']

@app.before_request
def csrf_protect():
    if request.method != 'POST':
        return None
    expected = session.get('csrf_token')
    if not expected:
        return abort(403, 'CSRF validation failed')
    supplied = (request.form.get('csrf_token')
                or (request.get_json(silent=True) or {}).get('csrf_token')
                or request.headers.get('X-CSRF-Token'))
    if not supplied or not secrets.compare_digest(supplied, expected):
        return abort(403, 'CSRF validation failed')
    return None

@app.context_processor
def inject_csrf():
    return {'csrf_token': ensure_csrf_token()}

# init DB tables on import (so gunicorn picks it up)
conn = sqlite3.connect(DB)
conn.execute('''CREATE TABLE IF NOT EXISTS items (
    id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, category TEXT NOT NULL,
    quantity INTEGER NOT NULL, price REAL,
    group_name TEXT, ram TEXT, rom TEXT, cost_price REAL, product_code TEXT, barcode TEXT)''')
conn.execute('''CREATE TABLE IF NOT EXISTS sales (
    id INTEGER PRIMARY KEY AUTOINCREMENT, item_id INTEGER NOT NULL, category TEXT NOT NULL,
    quantity INTEGER NOT NULL, timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
    receipt_id INTEGER, returned INTEGER DEFAULT 0, returned_at DATETIME,
    return_reason TEXT, created_by TEXT,
    imei TEXT, sellout_price TEXT, delivery_fee TEXT, special_note TEXT)''')
conn.execute('''CREATE TABLE IF NOT EXISTS customers (
    id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, phone TEXT,
    credit REAL DEFAULT 0, location TEXT, phone2 TEXT, phone3 TEXT)''')
conn.execute('''CREATE TABLE IF NOT EXISTS transactions (
    id INTEGER PRIMARY KEY AUTOINCREMENT, customer_id INTEGER NOT NULL,
    amount REAL NOT NULL, description TEXT,
    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP)''')
conn.execute('''CREATE TABLE IF NOT EXISTS brands (
    id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL,
    emoji TEXT NOT NULL DEFAULT '📦', color TEXT NOT NULL DEFAULT '⚪')''')
conn.execute('''CREATE TABLE IF NOT EXISTS product_groups (
    id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL,
    emoji TEXT NOT NULL DEFAULT '📦', color TEXT NOT NULL DEFAULT '⚪')''')
conn.execute('''CREATE TABLE IF NOT EXISTS staff (
    id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL,
    role TEXT DEFAULT 'staff', pin TEXT NOT NULL)''')
conn.execute('''CREATE TABLE IF NOT EXISTS customer_sales (
    id INTEGER PRIMARY KEY AUTOINCREMENT, customer_id INTEGER NOT NULL,
    item_id INTEGER NOT NULL, quantity INTEGER DEFAULT 1,
    price_paid REAL, note TEXT, sale_date TEXT,
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP, created_by TEXT)''')
conn.execute('''CREATE TABLE IF NOT EXISTS payment_schedules (
    id INTEGER PRIMARY KEY AUTOINCREMENT, customer_id INTEGER NOT NULL,
    transaction_id INTEGER, due_date TEXT NOT NULL, amount REAL NOT NULL,
    status TEXT DEFAULT 'pending', notes TEXT, created_by TEXT)''')
conn.execute('''CREATE TABLE IF NOT EXISTS item_imeis (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id INTEGER NOT NULL, imei TEXT NOT NULL UNIQUE, sold INTEGER DEFAULT 0)''')
conn.execute('''CREATE TABLE IF NOT EXISTS admin (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT UNIQUE NOT NULL,
    password_hash TEXT NOT NULL,
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
)''')
conn.execute('''CREATE TABLE IF NOT EXISTS bot_config (
    key TEXT PRIMARY KEY, value TEXT NOT NULL)''')
conn.execute("INSERT OR IGNORE INTO bot_config (key, value) VALUES ('unlock_pin', '123321')")
try:
    conn.execute("ALTER TABLE items ADD COLUMN barcode TEXT")
except:
    pass
conn.execute('''CREATE TABLE IF NOT EXISTS receipts (
    id INTEGER PRIMARY KEY AUTOINCREMENT, receipt_no TEXT UNIQUE NOT NULL,
    customer_id INTEGER, customer_name TEXT, total_amount REAL,
    discount REAL DEFAULT 0, payment_method TEXT DEFAULT 'cash',
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP)''')
try:
    conn.execute("ALTER TABLE sales ADD COLUMN receipt_id INTEGER")
except:
    pass
try:
    conn.execute("ALTER TABLE sales ADD COLUMN returned INTEGER DEFAULT 0")
except:
    pass
try:
    conn.execute("ALTER TABLE sales ADD COLUMN returned_at DATETIME")
except:
    pass
try:
    conn.execute("ALTER TABLE sales ADD COLUMN return_reason TEXT")
except:
    pass
try:
    conn.execute("ALTER TABLE sales ADD COLUMN created_by TEXT")
except:
    pass
try:
    conn.execute("ALTER TABLE payment_schedules ADD COLUMN created_by TEXT")
except:
    pass
try:
    conn.execute("ALTER TABLE customer_sales ADD COLUMN created_by TEXT")
except:
    pass
# Database indexes for performance
for idx in [
    "CREATE INDEX IF NOT EXISTS idx_sales_timestamp ON sales(timestamp)",
    "CREATE INDEX IF NOT EXISTS idx_sales_item_id ON sales(item_id)",
    "CREATE INDEX IF NOT EXISTS idx_items_category ON items(category)",
    "CREATE INDEX IF NOT EXISTS idx_items_name ON items(name)",
    "CREATE INDEX IF NOT EXISTS idx_stock_movements_created ON stock_movements(created_at)",
    "CREATE INDEX IF NOT EXISTS idx_stock_movements_item ON stock_movements(item_id)",
    "CREATE INDEX IF NOT EXISTS idx_payment_schedules_customer ON payment_schedules(customer_id)",
    "CREATE INDEX IF NOT EXISTS idx_customer_sales_customer ON customer_sales(customer_id)",
]:
    try:
        conn.execute(idx)
    except:
        pass
conn.commit()  # ensure no open transaction before WAL pragma
conn.execute("PRAGMA journal_mode=WAL")
conn.execute("PRAGMA foreign_keys=ON")
conn.commit()
conn.execute("INSERT OR IGNORE INTO bot_config (key, value) VALUES ('alert_bill_payment', '1')")
conn.execute("INSERT OR IGNORE INTO bot_config (key, value) VALUES ('alert_stock_runout', '1')")
conn.execute("INSERT OR IGNORE INTO bot_config (key, value) VALUES ('alert_import', '1')")
conn.execute('''CREATE TABLE IF NOT EXISTS notification_users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    chat_id INTEGER UNIQUE NOT NULL,
    name TEXT DEFAULT '',
    stock_alerts INTEGER DEFAULT 1,
    bill_alerts INTEGER DEFAULT 1,
    import_alerts INTEGER DEFAULT 1,
    active INTEGER DEFAULT 1,
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP)''')
# Migrate existing alert_chat_ids into notification_users
existing = conn.execute("SELECT value FROM bot_config WHERE key='alert_chat_ids'").fetchone()
if existing:
    ids = [int(x.strip()) for x in existing[0].split(',') if x.strip().isdigit()]
    for cid in ids:
        conn.execute("INSERT OR IGNORE INTO notification_users (chat_id) VALUES (?)", (cid,))
conn.execute('''CREATE TABLE IF NOT EXISTS stock_movements (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id INTEGER, item_name TEXT, item_category TEXT,
    change_type TEXT NOT NULL, qty_before INTEGER DEFAULT 0,
    qty_change INTEGER DEFAULT 0, qty_after INTEGER DEFAULT 0,
    price REAL, reference TEXT, created_by TEXT,
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP)''')
conn.commit()
conn.close()

DT = {
    'nav_dashboard': {'en': 'Dashboard', 'km': 'ផ្ទាំងគ្រប់គ្រង'},
    'nav_inventory': {'en': 'Inventory', 'km': 'ស្តុក'},
    'nav_customers': {'en': 'Customers', 'km': 'អតិថិជន'},
    'nav_sales': {'en': 'Sales', 'km': 'លក់'},
    'nav_payments': {'en': 'Payments', 'km': 'ការទូទាត់'},
    'nav_brands': {'en': 'Brands', 'km': 'ម៉ាក'},
    'nav_groups': {'en': 'Groups', 'km': 'ក្រុម'},
    'nav_staff': {'en': 'Staff', 'km': 'បុគ្គលិក'},
    'nav_exports': {'en': 'Exports', 'km': 'នាំចេញ'},
    'sellout': {'en': 'Sellout', 'km': 'លក់ចេញ'},
    'nav_change_password': {'en': 'Change Password', 'km': 'ប្តូរពាក្យសម្ងាត់'},
    'nav_logout': {'en': 'Logout', 'km': 'ចាកចេញ'},
    'title_dashboard': {'en': 'Dashboard', 'km': 'ផ្ទាំងគ្រប់គ្រង'},
    'title_inventory': {'en': 'Inventory', 'km': 'ស្តុក'},
    'title_customers': {'en': 'Customers', 'km': 'អតិថិជន'},
    'title_sales': {'en': 'Sales Report', 'km': 'របាយការណ៍លក់'},
    'title_payments': {'en': 'Payment Schedules', 'km': 'កាលវិភាគទូទាត់'},
    'title_brands': {'en': 'Brands', 'km': 'ម៉ាក'},
    'title_groups': {'en': 'Product Groups', 'km': 'ក្រុមផលិតផល'},
    'title_staff': {'en': 'Staff Management', 'km': 'គ្រប់គ្រងបុគ្គលិក'},
    'title_exports': {'en': 'Exported Files', 'km': 'ឯកសារនាំចេញ'},
    'total_items': {'en': 'Total Items', 'km': 'ទំនិញសរុប'},
    'stock_value': {'en': 'Stock Value (Cost)', 'km': 'តម្លៃស្តុក (ដើម)'},
    'customers_count': {'en': 'Customers', 'km': 'អតិថិជន'},
    'total_debt': {'en': 'Total Outstanding Debt', 'km': 'បំណុលសរុប'},
    'units_stock': {'en': 'units in stock', 'km': 'ដុំក្នុងស្តុក'},
    'retail_val': {'en': 'Retail', 'km': 'លក់រាយ'},
    'with_debt': {'en': 'with debt', 'km': 'មានបំណុល'},
    'pending_bills': {'en': 'pending bills', 'km': 'វិក្កយបត្រព្យួរ'},
    'recent_sales': {'en': 'Recent Sales', 'km': 'ការលក់ថ្មីៗ'},
    'upcoming_payments': {'en': 'Upcoming Payments', 'km': 'ការទូទាត់នាពេលខាងមុខ'},
    'item': {'en': 'Item', 'km': 'ទំនិញ'},
    'qty': {'en': 'Qty', 'km': 'បរិមាណ'},
    'revenue': {'en': 'Revenue', 'km': 'ចំណូល'},
    'date': {'en': 'Date', 'km': 'កាលបរិច្ឆេទ'},
    'customer': {'en': 'Customer', 'km': 'អតិថិជន'},
    'amount': {'en': 'Amount', 'km': 'ចំនួន'},
    'due': {'en': 'Due', 'km': 'កំណត់'},
    'status': {'en': 'Status', 'km': 'ស្ថានភាព'},
    'actions': {'en': 'Actions', 'km': 'សកម្មភាព'},
    'name': {'en': 'Name', 'km': 'ឈ្មោះ'},
    'brand': {'en': 'Brand', 'km': 'ម៉ាក'},
    'group': {'en': 'Group', 'km': 'ក្រុម'},
    'price': {'en': 'Price', 'km': 'តម្លៃ'},
    'cost': {'en': 'Cost', 'km': 'ថ្លៃដើម'},
    'margin': {'en': 'Margin', 'km': 'ប្រាក់ចំណេញ'},
    'stock': {'en': 'Stock', 'km': 'ស្តុក'},
    'value': {'en': 'Value', 'km': 'តម្លៃ'},
    'phone': {'en': 'Phone', 'km': 'ទូរស័ព្ទ'},
    'location': {'en': 'Location', 'km': 'ទីតាំង'},
    'debt': {'en': 'Debt', 'km': 'បំណុល'},
    'bills': {'en': 'Bills', 'km': 'វិក្កយបត្រ'},
    'profit': {'en': 'Profit', 'km': 'ប្រាក់ចំណេញ'},
    'imei': {'en': 'IMEI', 'km': 'IMEI'},
    'role': {'en': 'Role', 'km': 'តួនាទី'},
    'pin': {'en': 'PIN', 'km': 'PIN'},
    'search': {'en': 'Search...', 'km': 'ស្វែងរក...'},
    'all_groups': {'en': 'All Groups', 'km': 'ក្រុមទាំងអស់'},
    'all_brands': {'en': 'All Brands', 'km': 'ម៉ាកទាំងអស់'},
    'excel': {'en': 'Excel', 'km': 'Excel'},
    'new': {'en': 'New', 'km': 'ថ្មី'},
    'back': {'en': 'Back', 'km': 'ត្រឡប់'},
    'menu': {'en': 'Menu', 'km': 'ម៉ឺនុយ'},
    'login_title': {'en': 'Sign In', 'km': 'ចូល'},
    'login_btn': {'en': 'Sign In', 'km': 'ចូល'},
    'register_title': {'en': 'Create Admin Account', 'km': 'បង្កើតគណនីអ្នកគ្រប់គ្រង'},
    'register_btn': {'en': 'Create Account', 'km': 'បង្កើតគណនី'},
    'change_pw_title': {'en': 'Change Password', 'km': 'ប្តូរពាក្យសម្ងាត់'},
    'change_pw_btn': {'en': 'Change Password', 'km': 'ប្តូរពាក្យសម្ងាត់'},
    'username': {'en': 'Username', 'km': 'ឈ្មោះអ្នកប្រើ'},
    'password': {'en': 'Password', 'km': 'ពាក្យសម្ងាត់'},
    'confirm_pw': {'en': 'Confirm Password', 'km': 'បញ្ជាក់ពាក្យសម្ងាត់'},
    'current_pw': {'en': 'Current Password', 'km': 'ពាក្យសម្ងាត់បច្ចុប្បន្ន'},
    'new_pw': {'en': 'New Password', 'km': 'ពាក្យសម្ងាត់ថ្មី'},
    'no_data': {'en': 'No data available', 'km': 'គ្មានទិន្នន័យ'},
    'no_sales': {'en': 'No sales in this period', 'km': 'គ្មានការលក់ក្នុងរយៈពេលនេះ'},
    'no_items': {'en': 'No items found', 'km': 'រកមិនឃើញទំនិញ'},
    'no_customers': {'en': 'No customers', 'km': 'គ្មានអតិថិជន'},
    'no_exports': {'en': 'No exports yet', 'km': 'មិនទាន់មានការនាំចេញទេ'},
    'nav_settings': {'en': 'Settings', 'km': 'ការកំណត់'},
    'settings_title': {'en': 'Settings', 'km': 'ការកំណត់'},
    'bot_pin': {'en': 'Bot Unlock PIN', 'km': 'PIN ដោះសោ Bot'},
    'new_pin': {'en': 'New PIN', 'km': 'PIN ថ្មី'},
    'pin_updated': {'en': 'Bot PIN updated successfully', 'km': 'PIN ត្រូវបានធ្វើបច្ចុប្បន្នភាព'},
    'pin_mismatch': {'en': 'PINs do not match', 'km': 'PIN មិនត្រូវគ្នា'},
    'pin_too_short': {'en': 'PIN must be at least 4 characters', 'km': 'PIN ត្រូវមានយ៉ាងហោចណាស់ ៤ តួ'},
    'save_pin': {'en': 'Save PIN', 'km': 'រក្សាទុក PIN'},
    'nav_categories': {'en': 'Categories', 'km': 'ប្រភេទ'},
    'title_categories': {'en': 'Inventory Categories', 'km': 'ប្រភេទស្តុក'},
    'categories': {'en': 'Categories', 'km': 'ប្រភេទ'},
    'add_product': {'en': 'Add Product', 'km': 'បន្ថែមទំនិញ'},
    'barcode': {'en': 'Barcode', 'km': 'Barcode'},
    'category': {'en': 'Category', 'km': 'ប្រភេទ'},
    'items_count': {'en': 'Items', 'km': 'ទំនិញ'},
    'prod_in_cat': {'en': 'products in', 'km': 'ទំនិញក្នុង'},
    'no_category_items': {'en': 'No products in this category', 'km': 'គ្មានទំនិញក្នុងប្រភេទនេះទេ'},
}

def _d(key, lang=None):
    if lang is None:
        lang = session.get('lang', 'en')
    entry = DT.get(key, {})
    if lang in entry:
        return entry[lang]
    return entry.get('en', key)

def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def get_config(key, default=''):
    conn = get_db()
    cur = conn.execute("SELECT value FROM bot_config WHERE key=?", (key,))
    r = cur.fetchone()
    conn.close()
    return r['value'] if r else default

def log_stock_movement(conn, item_id, change_type, qty_change, qty_before=None, qty_after=None, price=None, reference='', created_by=None):
    try:
        cur = conn.execute("SELECT name, category, quantity FROM items WHERE id = ?", (item_id,))
        item = cur.fetchone()
        if not item:
            return
        name = item['name']
        cat = item['category']
        if qty_before is None:
            qty_before = item['quantity']
        if qty_after is None:
            qty_after = item['quantity'] + qty_change
        user = created_by or session.get('username', 'web')
        conn.execute(
            "INSERT INTO stock_movements (item_id, item_name, item_category, change_type, qty_before, qty_change, qty_after, price, reference, created_by) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (item_id, name, cat, change_type, qty_before, qty_change, qty_after, price, reference, user)
        )
    except Exception:
        pass

def export_excel(filename, headers, rows, sheet_name='Sheet1'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    hfont = Font(bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = Alignment(horizontal='center')
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    for col in ws.columns:
        ml = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 3, 45)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filepath = os.path.join(EXPORTS_DIR, f'{ts}_{filename}')
    wb.save(filepath)
    return filepath

@app.context_processor
def inject_now():
    staff_list = []
    try:
        conn = sqlite3.connect(DB)
        cur = conn.execute("SELECT id, name, role FROM staff ORDER BY name")
        staff_list = cur.fetchall()
        conn.close()
    except:
        pass
    return {'now': datetime.now(), '_d': _d, 'lang': lambda: session.get('lang', 'en'),
            'prices_unlocked': lambda: session.get('prices_unlocked', False),
            'staff_list': staff_list, 'staff_name': lambda: session.get('staff_name', '')}

@app.errorhandler(404)
def not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def server_error(e):
    return render_template('500.html'), 500



@app.route('/lang/<code>')
def set_lang(code):
    if code in ('en', 'km'):
        session['lang'] = code
    return redirect(request.referrer or url_for('index'))

@app.route('/set-theme', methods=['POST'])
def set_theme():
    theme = request.form.get('theme', 'light')
    if theme in ('light', 'dark'):
        session['theme'] = theme
    return ('', 204)

def login_required(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        if 'admin_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapper

@app.route('/')
def index():
    if 'admin_id' in session:
        return redirect(url_for('dashboard'))
    conn = get_db()
    cur = conn.execute("SELECT COUNT(*) FROM admin")
    has_admin = cur.fetchone()[0] > 0
    conn.close()
    if not has_admin:
        return redirect(url_for('register'))
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    conn = get_db()
    cur = conn.execute("SELECT COUNT(*) FROM admin")
    has_admin = cur.fetchone()[0] > 0
    if has_admin:
        conn.close()
        return redirect(url_for('login'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm = request.form.get('confirm', '')
        if not username or not password:
            flash('Username and password required', 'danger')
        elif password != confirm:
            flash('Passwords do not match', 'danger')
        elif len(password) < 4:
            flash('Password must be at least 4 characters', 'danger')
        else:
            try:
                conn.execute("INSERT INTO admin (username, password_hash) VALUES (?, ?)",
                             (username, generate_password_hash(password)))
                conn.commit()
                flash('Admin account created! Please log in.', 'success')
                conn.close()
                return redirect(url_for('login'))
            except sqlite3.IntegrityError:
                flash('Username already exists', 'danger')
    conn.close()
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'admin_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        conn = get_db()
        cur = conn.execute("SELECT * FROM admin WHERE username = ?", (username,))
        admin = cur.fetchone()
        conn.close()
        if admin and check_password_hash(admin['password_hash'], password):
            session['admin_id'] = admin['id']
            session['admin_username'] = admin['username']
            flash('Logged in successfully', 'success')
            return redirect(url_for('dashboard'))
        flash('Invalid username or password', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out', 'info')
    return redirect(url_for('login'))

@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    conn = get_db()
    if request.method == 'POST':
        if 'new_password' in request.form:
            current = request.form.get('current', '')
            new_pass = request.form.get('new_password', '')
            confirm = request.form.get('confirm', '')
            cur = conn.execute("SELECT * FROM admin WHERE id = ?", (session['admin_id'],))
            admin = cur.fetchone()
            if not admin or not check_password_hash(admin['password_hash'], current):
                flash('Current password is incorrect', 'danger')
            elif new_pass != confirm:
                flash('New passwords do not match', 'danger')
            elif len(new_pass) < 4:
                flash('Password must be at least 4 characters', 'danger')
            else:
                conn.execute("UPDATE admin SET password_hash = ? WHERE id = ?",
                             (generate_password_hash(new_pass), session['admin_id']))
                conn.commit()
                flash('Password changed successfully', 'success')
        elif 'new_pin' in request.form:
            new_pin = request.form.get('new_pin', '')
            confirm_pin = request.form.get('confirm_pin', '')
            if new_pin != confirm_pin:
                flash('PINs do not match', 'danger')
            elif len(new_pin) < 4:
                flash('PIN must be at least 4 characters', 'danger')
            else:
                conn.execute("UPDATE bot_config SET value = ? WHERE key = 'unlock_pin'", (new_pin,))
                conn.commit()
                flash('Bot PIN updated successfully', 'success')
        elif 'stock_threshold' in request.form:
            val = request.form.get('stock_threshold', '').strip()
            if val.isdigit() and int(val) > 0:
                conn.execute("UPDATE bot_config SET value = ? WHERE key = 'stock_alert_threshold'", (val,))
                conn.commit()
                flash('Stock alert threshold updated', 'success')
            else:
                flash('Enter a valid positive number', 'danger')
        elif 'alert_chat_ids' in request.form:
            ids = request.form.get('alert_chat_ids', '').strip()
            if ids and all(x.strip().isdigit() for x in ids.split(',') if x.strip()):
                conn.execute("UPDATE bot_config SET value = ? WHERE key = 'alert_chat_ids'", (ids,))
                conn.commit()
                flash('Alert recipients updated', 'success')
            else:
                flash('Enter comma-separated numeric Telegram IDs', 'danger')
        elif 'alert_interval' in request.form:
            val = request.form.get('alert_interval', '').strip()
            try:
                v = float(val)
                if v >= 0:
                    conn.execute("UPDATE bot_config SET value = ? WHERE key = 'alert_interval_hours'", (str(v),))
                    conn.commit()
                    flash('Alert interval updated', 'success')
                else:
                    flash('Enter 0 or a positive number', 'danger')
            except ValueError:
                flash('Enter a valid number (hours)', 'danger')
        elif 'bill_alert_days' in request.form:
            val = request.form.get('bill_alert_days', '').strip()
            if val.isdigit() and int(val) > 0:
                conn.execute("UPDATE bot_config SET value = ? WHERE key = 'bill_alert_days'", (val,))
                conn.commit()
                flash('Bill alert days updated', 'success')
            else:
                flash('Enter a valid positive number', 'danger')
        elif 'bot_user_level' in request.form:
            level = request.form.get('bot_user_level', '').strip()
            if level in ('admin', 'staff'):
                conn.execute("UPDATE bot_config SET value = ? WHERE key = 'bot_user_level'", (level,))
                conn.commit()
                flash('Bot user level updated', 'success')
            else:
                flash('Invalid level', 'danger')
        elif 'action' in request.form and request.form['action'] == 'toggle_startup':
            # Toggle auto-startup shortcut
            startup_path = os.path.join(os.environ['APPDATA'], 'Microsoft', 'Windows', 'Start Menu', 'Programs', 'Startup', 'Inventory Bot.lnk')
            if os.path.exists(startup_path):
                os.remove(startup_path)
                flash('Auto-startup disabled', 'success')
            else:
                import subprocess, sys
                python_exe = os.path.join(os.path.dirname(sys.executable), 'python.exe') if sys.executable else sys.executable
                launcher = os.path.join(os.path.dirname(__file__), 'launcher.py')
                # Backup: also try the exe on Desktop
                desktop_exe = os.path.join(os.environ['USERPROFILE'], 'Desktop', 'Inventory.exe')
                target = f'"{desktop_exe}"' if os.path.exists(desktop_exe) else f'"{python_exe}" "{launcher}"'
                try:
                    ws = __import__('win32com', fromlist=['client']).client.Dispatch('WScript.Shell')
                except ImportError:
                    flash('Auto-startup requires pywin32 (pip install pywin32)', 'danger')
                    return redirect(url_for('settings'))
                sc = ws.CreateShortcut(startup_path)
                import shlex
                parts = shlex.split(target)
                sc.TargetPath = parts[0] if parts else ''
                sc.Arguments = ' '.join(parts[1:])
                sc.WorkingDirectory = os.path.dirname(__file__)
                sc.WindowStyle = 7
                sc.Description = 'Inventory Bot Auto-Start'
                sc.Save()
                flash('Auto-startup enabled', 'success')
        elif 'action' in request.form and request.form['action'] == 'test_notification':
            target_ids_str = request.form.get('test_chat_id', '').strip()
            if target_ids_str and all(x.strip().isdigit() for x in target_ids_str.split(',') if x.strip()):
                flash(f'Test notification will be sent on next bot restart to: {target_ids_str}', 'info')
            else:
                flash('Enter a valid Telegram ID', 'danger')
        elif 'action' in request.form and request.form['action'] == 'add_notification_user':
            chat_id = request.form.get('chat_id', '').strip()
            name = request.form.get('name', '').strip()
            if chat_id.isdigit():
                conn.execute("INSERT OR IGNORE INTO notification_users (chat_id, name) VALUES (?, ?)", (int(chat_id), name))
                conn.commit()
                flash('Notification user added', 'success')
            else:
                flash('Enter a valid numeric Telegram ID', 'danger')
        elif 'action' in request.form and request.form['action'] == 'toggle_notification_alert':
            try:
                uid = int(request.form.get('user_id', 0))
                alert_type = request.form.get('alert_type', '')
                if alert_type in ('stock_alerts', 'bill_alerts', 'import_alerts'):
                    cur = conn.execute(f"SELECT {alert_type} FROM notification_users WHERE id=?", (uid,))
                    row = cur.fetchone()
                    if row:
                        new_val = 0 if row[0] else 1
                        conn.execute(f"UPDATE notification_users SET {alert_type}=? WHERE id=?", (new_val, uid))
                        conn.commit()
                        flash('Alert preference updated', 'success')
            except Exception:
                flash('Error updating preference', 'danger')
        elif 'action' in request.form and request.form['action'] == 'delete_notification_user':
            try:
                uid = int(request.form.get('user_id', 0))
                conn.execute("DELETE FROM notification_users WHERE id=?", (uid,))
                conn.commit()
                flash('Notification user removed', 'success')
            except Exception:
                flash('Error removing user', 'danger')
    # Read all settings
    current_threshold = get_config('stock_alert_threshold', '5')
    current_ids = get_config('alert_chat_ids', '7185846273')
    current_interval = get_config('alert_interval_hours', '0')
    bill_alert_days = get_config('bill_alert_days', '7')
    bot_user_level = get_config('bot_user_level', 'admin')
    notif_users = conn.execute("SELECT * FROM notification_users ORDER BY active DESC, id ASC").fetchall()
    startup_path = os.path.join(os.environ.get('APPDATA', ''), 'Microsoft', 'Windows', 'Start Menu', 'Programs', 'Startup', 'Inventory Bot.lnk')
    auto_startup = 'yes' if os.path.exists(startup_path) else 'no'
    return render_template('settings.html',
        stock_threshold=current_threshold, alert_chat_ids=current_ids,
        alert_interval=current_interval, bill_alert_days=bill_alert_days,
        bot_user_level=bot_user_level, auto_startup=auto_startup,
        notif_users=notif_users)

@app.route('/unlock-prices', methods=['POST'])
@login_required
def unlock_prices():
    pin = request.form.get('pin', '')
    conn = get_db()
    cur = conn.execute("SELECT value FROM bot_config WHERE key = 'unlock_pin'")
    row = cur.fetchone()
    conn.close()
    expected = row['value'] if row else '123321'
    if pin == expected:
        session['prices_unlocked'] = True
        return {'ok': True, 'unlocked': True}
    return {'ok': False, 'msg': 'Wrong PIN'}, 403

@app.route('/lock-prices')
@login_required
def lock_prices():
    session.pop('prices_unlocked', None)
    return {'ok': True, 'unlocked': False}

@app.route('/set-staff', methods=['POST'])
@login_required
def set_staff():
    staff_name = request.form.get('staff_name', '').strip()
    if staff_name:
        session['staff_name'] = staff_name
    else:
        session.pop('staff_name', None)
    flash(f'Active staff: {staff_name or "None"}', 'info')
    return redirect(request.referrer or url_for('dashboard'))

@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()
    stats = {}

    # Date filter for sellout
    from_date = request.args.get('from', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    to_date = request.args.get('to', datetime.now().strftime('%Y-%m-%d'))

    cur = conn.execute("SELECT COUNT(*) FROM items")
    stats['total_items'] = cur.fetchone()[0]
    cur = conn.execute("SELECT COALESCE(SUM(quantity),0) FROM items")
    stats['total_stock'] = cur.fetchone()[0]
    cur = conn.execute("SELECT COALESCE(SUM(quantity * cost_price),0) FROM items")
    stats['total_investment'] = cur.fetchone()[0]
    cur = conn.execute("SELECT COALESCE(SUM(quantity * price),0) FROM items")
    stats['total_retail'] = cur.fetchone()[0]
    cur = conn.execute("SELECT COUNT(*) FROM customers")
    stats['total_customers'] = cur.fetchone()[0]
    cur = conn.execute("SELECT COALESCE(SUM(amount),0) FROM payment_schedules WHERE status IN ('pending','partial')")
    stats['total_debt'] = cur.fetchone()[0]
    cur = conn.execute("SELECT COUNT(*) FROM payment_schedules WHERE status IN ('pending','partial')")
    stats['pending_bills'] = cur.fetchone()[0]

    cur = conn.execute('''SELECT s.id, i.name as item_name, s.quantity,
        s.timestamp, COALESCE(CAST(NULLIF(s.sellout_price,'') AS REAL), i.price) as price, i.cost_price
        FROM sales s JOIN items i ON s.item_id = i.id
        ORDER BY s.timestamp DESC LIMIT 10''')
    stats['recent_sales'] = cur.fetchall()

    cur = conn.execute('''SELECT c.name as customer_name, ps.amount, ps.due_date, ps.status, ps.id
        FROM payment_schedules ps JOIN customers c ON ps.customer_id = c.id
        WHERE ps.status IN ('pending','partial')
        ORDER BY ps.due_date ASC LIMIT 10''')
    stats['upcoming_payments'] = cur.fetchall()

    stats['stock_value'] = stats['total_investment']

    # Bill payment chart data — paid vs pending by month
    cur = conn.execute('''SELECT strftime('%Y-%m', due_date) as ym,
        SUM(CASE WHEN status='paid' THEN amount ELSE 0 END) as paid,
        SUM(CASE WHEN status IN ('pending','partial') THEN amount ELSE 0 END) as pending
        FROM payment_schedules GROUP BY ym ORDER BY ym LIMIT 12''')
    bill_chart = cur.fetchall()
    stats['bill_chart_labels'] = [r['ym'] for r in bill_chart]
    stats['bill_chart_paid'] = [r['paid'] for r in bill_chart]
    stats['bill_chart_pending'] = [r['pending'] for r in bill_chart]

    # Stock by brand chart
    cur = conn.execute('''SELECT COALESCE(b.name, i.category) as brand_name, SUM(i.quantity) as total_qty
        FROM items i LEFT JOIN brands b ON i.category = b.name
        GROUP BY brand_name ORDER BY total_qty DESC LIMIT 15''')
    brand_stock = cur.fetchall()
    stats['brand_stock_labels'] = [r['brand_name'] or 'Other' for r in brand_stock]
    stats['brand_stock_data'] = [r['total_qty'] for r in brand_stock]

    # Sellout chart by date range
    cur = conn.execute('''SELECT DATE(s.timestamp) as day, SUM(s.quantity) as qty,
        COALESCE(SUM(CAST(NULLIF(s.sellout_price,'') AS REAL)), SUM(i.price * s.quantity)) as rev
        FROM sales s JOIN items i ON s.item_id = i.id
        WHERE DATE(s.timestamp) >= ? AND DATE(s.timestamp) <= ?
        GROUP BY day ORDER BY day''', (from_date, to_date))
    sellout_data = cur.fetchall()
    stats['sellout_dates'] = [r['day'] for r in sellout_data]
    stats['sellout_qty'] = [r['qty'] for r in sellout_data]
    stats['sellout_rev'] = [r['rev'] for r in sellout_data]

    # Alert preferences
    stats['alert_bill'] = get_config('alert_bill_payment', '1')
    stats['alert_stock'] = get_config('alert_stock_runout', '1')
    stats['alert_import'] = get_config('alert_import', '1')

    # Low-stock items (qty <= threshold, default 5)
    threshold = int(get_config('stock_alert_threshold', '5'))
    cur = conn.execute("SELECT i.name, i.category, i.quantity, b.emoji as brand_emoji FROM items i LEFT JOIN brands b ON i.category = b.name WHERE i.quantity > 0 AND i.quantity <= ? ORDER BY i.quantity ASC LIMIT 20", (threshold,))
    stats['low_stock_items'] = cur.fetchall()

    stats['from_date'] = from_date
    stats['to_date'] = to_date

    conn.close()
    return render_template('dashboard.html', **stats)

@app.route('/dashboard/alert-prefs', methods=['POST'])
@login_required
def save_alert_prefs():
    bill = '1' if request.form.get('alert_bill') else '0'
    stock = '1' if request.form.get('alert_stock') else '0'
    imp = '1' if request.form.get('alert_import') else '0'
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO bot_config (key, value) VALUES ('alert_bill_payment', ?)", (bill,))
    conn.execute("INSERT OR REPLACE INTO bot_config (key, value) VALUES ('alert_stock_runout', ?)", (stock,))
    conn.execute("INSERT OR REPLACE INTO bot_config (key, value) VALUES ('alert_import', ?)", (imp,))
    conn.commit()
    conn.close()
    flash('Alert preferences saved', 'success')
    return redirect(url_for('dashboard'))

@app.route('/inventory')
@login_required
def inventory():
    conn = get_db()
    brand = request.args.get('brand', '')
    group = request.args.get('group', '')
    search = request.args.get('search', '')
    query = '''SELECT i.*, b.emoji as brand_emoji, b.color as brand_color, b.name as brand_name,
               pg.emoji as group_emoji, pg.color as group_color
               FROM items i
               LEFT JOIN brands b ON i.category = b.name
               LEFT JOIN product_groups pg ON i.group_name = pg.name
               WHERE 1=1'''
    params = []
    if brand:
        query += ' AND i.category = ?'
        params.append(brand)
    if group:
        query += ' AND i.group_name = ?'
        params.append(group)
    if search:
        query += ' AND (i.name LIKE ? OR i.category LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%'])
    query += ' ORDER BY i.category, i.group_name, i.name'
    cur = conn.execute(query, params)
    items = cur.fetchall()
    # Group items by brand → group
    grouped = {}
    brand_counts = {}
    total_count = 0
    total_value = 0
    for item in items:
        b = item['brand_name'] or item['category'] or 'Other'
        g = item['group_name'] or 'General'
        if b not in grouped:
            grouped[b] = {}
            brand_counts[b] = 0
        if g not in grouped[b]:
            grouped[b][g] = []
        grouped[b][g].append(item)
        brand_counts[b] += 1
        total_count += item['quantity']
        total_value += (item['cost_price'] or 0) * item['quantity']
    cur = conn.execute("SELECT name FROM brands ORDER BY name")
    brands = cur.fetchall()
    cur = conn.execute("SELECT name FROM product_groups ORDER BY name")
    groups = cur.fetchall()
    conn.close()
    return render_template('inventory.html', grouped=grouped, brands=brands, groups=groups,
                           brand=brand, group=group, search=search,
                           total_count=total_count, total_value=total_value,
                           brand_counts=brand_counts)

@app.route('/inventory/add-stock', methods=['POST'])
@login_required
def add_stock():
    item_id = request.form.get('item_id')
    qty = request.form.get('qty', '0')
    cost = request.form.get('cost_price', '')
    try:
        qty = int(qty)
        if qty <= 0:
            flash('Quantity must be positive', 'danger')
            return redirect(url_for('inventory'))
        conn = get_db()
        cur = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,))
        item = cur.fetchone()
        if not item:
            flash('Item not found', 'danger')
        else:
            new_qty = item['quantity'] + qty
            if cost and cost.strip():
                cost_val = float(cost.replace('$','').replace(',',''))
                if item['cost_price'] and item['quantity'] > 0:
                    total_cost = (item['cost_price'] * item['quantity']) + (cost_val * qty)
                    avg_cost = total_cost / new_qty
                else:
                    avg_cost = cost_val
            else:
                avg_cost = item['cost_price']
                cost_val = item['cost_price']
            conn.execute("UPDATE items SET quantity = ?, cost_price = ? WHERE id = ?",
                         (new_qty, avg_cost, item_id))
            log_stock_movement(conn, item_id, 'addition', qty, item['quantity'], new_qty, cost_val or item['cost_price'], created_by=session.get('staff_name', '')) 
            conn.commit()
            flash(f'Added {qty} stock. New qty: {new_qty}', 'success')
        conn.close()
    except (ValueError, TypeError):
        flash('Invalid input', 'danger')
    return redirect(url_for('inventory'))

@app.route('/inventory/add-item', methods=['POST'])
@login_required
def add_item():
    name = request.form.get('name', '').strip()
    brand = request.form.get('brand', '').strip()
    group = request.form.get('group', '').strip()
    qty = request.form.get('qty', '0')
    price = request.form.get('price', '')
    cost = request.form.get('cost_price', '')
    ram = request.form.get('ram', '').strip()
    rom = request.form.get('rom', '').strip()
    code = request.form.get('product_code', '').strip()
    barcode_val = request.form.get('barcode', '').strip()
    if not name:
        flash('Name is required', 'danger')
        return redirect(url_for('inventory'))
    if not brand:
        flash('Brand is required', 'danger')
        return redirect(url_for('inventory'))
    try:
        qty = int(qty)
    except:
        qty = 0
    try:
        price = float(price.replace('$','').replace(',','')) if price else None
    except:
        price = None
    try:
        cost = float(cost.replace('$','').replace(',','')) if cost else None
    except:
        cost = None
    if not barcode_val:
        import random
        barcode_val = f'BR-{random.randint(10000, 99999)}-{random.randint(100, 999)}'
    conn = get_db()
    conn.execute("INSERT INTO items (name,category,quantity,price,cost_price,ram,rom,group_name,product_code,barcode) VALUES (?,?,?,?,?,?,?,?,?,?)",
                 (name, brand, qty, price, cost, ram, rom, group, code, barcode_val))
    conn.commit()
    conn.close()
    flash(f'✅ Added {name}', 'success')
    return redirect(url_for('inventory'))

@app.route('/inventory/edit', methods=['POST'])
@login_required
def edit_item():
    item_id = request.form.get('item_id')
    price = request.form.get('price', '')
    name = request.form.get('name', '')
    ram = request.form.get('ram', '')
    rom = request.form.get('rom', '')
    code = request.form.get('product_code', '')
    group = request.form.get('group', '')
    cost_price = request.form.get('cost_price', '')
    conn = get_db()
    updates = []
    params = []
    if price:
        try:
            updates.append('price = ?')
            params.append(float(price.replace('$','').replace(',','')))
        except: pass
    if cost_price:
        try:
            updates.append('cost_price = ?')
            params.append(float(cost_price.replace('$','').replace(',','')))
        except: pass
    if name:
        updates.append('name = ?')
        params.append(name)
    if ram is not None:
        updates.append('ram = ?')
        params.append(ram)
    if rom is not None:
        updates.append('rom = ?')
        params.append(rom)
    if code is not None:
        updates.append('product_code = ?')
        params.append(code)
    if group:
        updates.append('group_name = ?')
        params.append(group)
    if updates:
        params.append(item_id)
        conn.execute(f"UPDATE items SET {', '.join(updates)} WHERE id = ?", params)
        conn.commit()
    conn.close()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return {'ok': True, 'msg': 'Item updated'}
    flash('Item updated', 'success')
    return redirect(url_for('inventory'))

@app.route('/inventory/delete/<int:item_id>', methods=['POST'])
@login_required
def delete_item(item_id):
    conn = get_db()
    cur = conn.execute("SELECT name, quantity, cost_price FROM items WHERE id = ?", (item_id,))
    item = cur.fetchone()
    if not item:
        conn.close()
        return {'ok': False, 'msg': 'Item not found'}
    cur = conn.execute("SELECT id FROM sales WHERE item_id = ? LIMIT 1", (item_id,))
    has_sales = cur.fetchone()
    if has_sales:
        conn.close()
        return {'ok': False, 'msg': 'Cannot delete: item has sales history'}
    conn.execute("DELETE FROM item_imeis WHERE item_id = ?", (item_id,))
    log_stock_movement(conn, item_id, 'deletion', -item['quantity'], item['quantity'], 0, item['cost_price'], f'deleted by {session.get("username","web")}')
    conn.execute("DELETE FROM items WHERE id = ?", (item_id,))
    conn.commit()
    conn.close()
    return {'ok': True, 'msg': f'"{item["name"]}" deleted'}

@app.route('/inventory/<int:item_id>', methods=['GET', 'POST'])
@login_required
def item_detail(item_id):
    conn = get_db()
    if request.method == 'POST':
        price = request.form.get('price', '')
        cost_price = request.form.get('cost_price', '')
        name = request.form.get('name', '')
        ram = request.form.get('ram', '')
        rom = request.form.get('rom', '')
        code = request.form.get('product_code', '')
        group = request.form.get('group', '')
        updates = []
        params = []
        if price:
            try:
                updates.append('price = ?')
                params.append(float(price.replace('$','').replace(',','')))
            except: pass
        if cost_price:
            try:
                updates.append('cost_price = ?')
                params.append(float(cost_price.replace('$','').replace(',','')))
            except: pass
        if name:
            updates.append('name = ?')
            params.append(name)
        if ram is not None:
            updates.append('ram = ?')
            params.append(ram)
        if rom is not None:
            updates.append('rom = ?')
            params.append(rom)
        if code is not None:
            updates.append('product_code = ?')
            params.append(code)
        if group:
            updates.append('group_name = ?')
            params.append(group)
        if updates:
            params.append(item_id)
            conn.execute(f"UPDATE items SET {', '.join(updates)} WHERE id = ?", params)
            conn.commit()
            flash('Item updated', 'success')
        return redirect(url_for('item_detail', item_id=item_id))
    cur = conn.execute('''SELECT i.*, b.emoji as brand_emoji, b.color as brand_color,
                          pg.emoji as group_emoji, pg.color as group_color
                          FROM items i
                          LEFT JOIN brands b ON i.category = b.name
                          LEFT JOIN product_groups pg ON i.group_name = pg.name
                          WHERE i.id = ?''', (item_id,))
    item = cur.fetchone()
    if not item:
        flash('Item not found', 'danger')
        return redirect(url_for('inventory'))
    cur = conn.execute('''SELECT s.id, s.quantity, s.timestamp, s.sellout_price, s.imei,
                          '' as customer_name, 0 as cid
                          FROM sales s
                          WHERE s.item_id = ?
                          ORDER BY s.timestamp DESC LIMIT 20''', (item_id,))
    sales_history = cur.fetchall()
    conn.close()
    return render_template('item_detail.html', item=item, sales_history=sales_history)

@app.route('/customers')
@login_required
def customers():
    conn = get_db()
    search = request.args.get('search', '')
    query = '''SELECT c.*,
               (SELECT COALESCE(SUM(amount),0) FROM payment_schedules
                WHERE customer_id = c.id AND status IN ('pending','partial')) as total_debt,
               (SELECT COUNT(*) FROM payment_schedules
                WHERE customer_id = c.id AND status IN ('pending','partial')) as pending_bills
               FROM customers c'''
    params = []
    if search:
        query += ' WHERE c.name LIKE ? OR c.phone LIKE ?'
        params.extend([f'%{search}%', f'%{search}%'])
    query += ' ORDER BY c.name'
    cur = conn.execute(query, params)
    customers = cur.fetchall()
    conn.close()
    return render_template('customers.html', customers=customers, search=search)

@app.route('/customers/<int:cust_id>')
@login_required
def customer_detail(cust_id):
    conn = get_db()
    cur = conn.execute("SELECT * FROM customers WHERE id = ?", (cust_id,))
    customer = cur.fetchone()
    if not customer:
        flash('Customer not found', 'danger')
        return redirect(url_for('customers'))
    cur = conn.execute('''SELECT * FROM payment_schedules
                          WHERE customer_id = ? ORDER BY due_date ASC''', (cust_id,))
    bills = cur.fetchall()
    cur = conn.execute('''SELECT cs.*, i.name as item_name, i.category as brand
                          FROM customer_sales cs
                          JOIN items i ON cs.item_id = i.id
                          WHERE cs.customer_id = ?
                          ORDER BY cs.created_at DESC LIMIT 50''', (cust_id,))
    buys = cur.fetchall()
    cur = conn.execute('''SELECT * FROM transactions
                          WHERE customer_id = ? ORDER BY timestamp DESC LIMIT 50''', (cust_id,))
    transactions = cur.fetchall()
    cur = conn.execute('''SELECT COALESCE(SUM(amount),0) FROM payment_schedules
                          WHERE customer_id = ? AND status IN ('pending','partial')''', (cust_id,))
    total_debt = cur.fetchone()[0]
    conn.close()
    return render_template('customer_detail.html', customer=customer, bills=bills,
                           buys=buys, transactions=transactions, total_debt=total_debt)

@app.route('/customers/<int:cust_id>/statement')
@login_required
def customer_statement(cust_id):
    conn = get_db()
    cur = conn.execute("SELECT * FROM customers WHERE id = ?", (cust_id,))
    customer = cur.fetchone()
    if not customer:
        conn.close()
        flash('Customer not found', 'danger')
        return redirect(url_for('customers'))
    cur = conn.execute("SELECT * FROM payment_schedules WHERE customer_id = ? ORDER BY due_date ASC", (cust_id,))
    bills = cur.fetchall()
    cur = conn.execute('''SELECT cs.*, i.name as item_name, i.category as brand
        FROM customer_sales cs JOIN items i ON cs.item_id = i.id
        WHERE cs.customer_id = ? ORDER BY cs.created_at DESC''', (cust_id,))
    buys = cur.fetchall()
    cur = conn.execute("SELECT * FROM transactions WHERE customer_id = ? ORDER BY timestamp ASC", (cust_id,))
    transactions = cur.fetchall()
    cur = conn.execute("SELECT COALESCE(SUM(amount),0) FROM payment_schedules WHERE customer_id = ? AND status IN ('pending','partial')", (cust_id,))
    total_debt = cur.fetchone()[0]
    conn.close()
    return render_template('customer_statement.html', customer=customer, bills=bills,
                           buys=buys, transactions=transactions, total_debt=total_debt, now=datetime.now())

@app.route('/customers/create', methods=['POST'])
@login_required
def create_customer():
    name = request.form.get('name', '').strip()
    phone = request.form.get('phone', '').strip()
    location = request.form.get('location', '').strip()
    phone2 = request.form.get('phone2', '').strip()
    phone3 = request.form.get('phone3', '').strip()
    if not name:
        flash('Customer name is required', 'danger')
        return redirect(url_for('customers'))
    conn = get_db()
    conn.execute("INSERT INTO customers (name, phone, location, phone2, phone3) VALUES (?,?,?,?,?)",
                 (name, phone or None, location or None, phone2 or None, phone3 or None))
    conn.commit()
    conn.close()
    flash(f'Customer "{name}" created', 'success')
    return redirect(url_for('customers'))

@app.route('/customers/<int:cust_id>/edit', methods=['POST'])
@login_required
def edit_customer(cust_id):
    name = request.form.get('name', '').strip()
    phone = request.form.get('phone', '').strip()
    location = request.form.get('location', '').strip()
    phone2 = request.form.get('phone2', '').strip()
    phone3 = request.form.get('phone3', '').strip()
    if not name:
        flash('Customer name is required', 'danger')
        return redirect(url_for('customer_detail', cust_id=cust_id))
    conn = get_db()
    conn.execute("UPDATE customers SET name=?, phone=?, location=?, phone2=?, phone3=? WHERE id=?",
                 (name, phone or None, location or None, phone2 or None, phone3 or None, cust_id))
    conn.commit()
    conn.close()
    flash('Customer updated', 'success')
    return redirect(url_for('customer_detail', cust_id=cust_id))

@app.route('/customers/<int:cust_id>/delete', methods=['POST'])
@login_required
def delete_customer(cust_id):
    conn = get_db()
    cur = conn.execute("SELECT name FROM customers WHERE id=?", (cust_id,))
    cust = cur.fetchone()
    if not cust:
        conn.close()
        flash('Customer not found', 'danger')
        return redirect(url_for('customers'))
    conn.execute("DELETE FROM payment_schedules WHERE customer_id=?", (cust_id,))
    conn.execute("DELETE FROM transactions WHERE customer_id=?", (cust_id,))
    conn.execute("DELETE FROM customer_sales WHERE customer_id=?", (cust_id,))
    conn.execute("DELETE FROM customers WHERE id=?", (cust_id,))
    conn.commit()
    conn.close()
    flash(f'Customer "{cust["name"]}" deleted', 'success')
    return redirect(url_for('customers'))

@app.route('/export/bills')
@login_required
def export_bills():
    conn = get_db()
    cur = conn.execute('''SELECT c.name, c.phone, ps.amount, ps.due_date, ps.status, ps.notes
                          FROM payment_schedules ps JOIN customers c ON ps.customer_id = c.id
                          ORDER BY ps.due_date ASC''')
    rows = cur.fetchall()
    conn.close()
    headers = ['Customer', 'Phone', 'Amount', 'Due Date', 'Status', 'Description']
    data = [[r[0], r[1], r[2], r[3], r[4], r[5]] for r in rows]
    fp = export_excel('bills_report.xlsx', headers, data)
    return send_file(fp, as_attachment=True, download_name='bills_report.xlsx')

@app.route('/sales')
@login_required
def sales():
    conn = get_db()
    days = request.args.get('days', '7')
    try:
        days = int(days)
    except:
        days = 7
    start = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    cur = conn.execute('''SELECT s.*, i.name as item_name, i.category as brand,
                          COALESCE(CAST(NULLIF(s.sellout_price,'') AS REAL), i.price) as price,
                          i.cost_price
                          FROM sales s
                          JOIN items i ON s.item_id = i.id
                          WHERE DATE(s.timestamp) >= ?
                          ORDER BY s.timestamp DESC''', (start,))
    sales = cur.fetchall()
    cur = conn.execute('''SELECT COALESCE(SUM(s.quantity * COALESCE(CAST(NULLIF(s.sellout_price,'') AS REAL), i.price)),0)
                          FROM sales s JOIN items i ON s.item_id = i.id
                          WHERE DATE(s.timestamp) >= ?''', (start,))
    total_revenue = cur.fetchone()[0]
    cur = conn.execute('''SELECT COALESCE(SUM(s.quantity * (COALESCE(CAST(NULLIF(s.sellout_price,'') AS REAL), i.price) - i.cost_price)),0)
                          FROM sales s JOIN items i ON s.item_id = i.id
                          WHERE DATE(s.timestamp) >= ? AND i.cost_price IS NOT NULL''', (start,))
    total_profit = cur.fetchone()[0]
    conn.close()
    return render_template('sales.html', sales=sales, total_revenue=total_revenue,
                           total_profit=total_profit, days=days)

@app.route('/schedules')
@login_required
def schedules():
    conn = get_db()
    status = request.args.get('status', 'all')
    query = '''SELECT ps.*, c.name as customer_name, c.phone
               FROM payment_schedules ps JOIN customers c ON ps.customer_id = c.id'''
    params = []
    if status != 'all':
        query += ' WHERE ps.status = ?'
        params.append(status)
    query += ' ORDER BY ps.due_date ASC'
    cur = conn.execute(query, params)
    schedules = cur.fetchall()
    conn.close()
    return render_template('schedules.html', schedules=schedules, status=status)

@app.route('/schedules/mark-paid', methods=['POST'])
@login_required
def mark_paid():
    schedule_id = request.form.get('schedule_id')
    conn = get_db()
    cur = conn.execute("SELECT amount, customer_id, status FROM payment_schedules WHERE id = ?", (schedule_id,))
    sched = cur.fetchone()
    if sched and sched['status'] != 'paid':
        conn.execute("UPDATE payment_schedules SET status = 'paid', created_by = ? WHERE id = ?", (session.get('staff_name', ''), schedule_id))
        conn.execute("INSERT INTO transactions (customer_id, amount, description) VALUES (?,?,?)",
                     (sched['customer_id'], -sched['amount'], f'Payment for bill #{schedule_id}'))
        conn.execute("UPDATE customers SET credit = credit - ? WHERE id = ?", (sched['amount'], sched['customer_id']))
    conn.commit()
    conn.close()
    flash('Marked as paid', 'success')
    return redirect(url_for('schedules'))

@app.route('/brands')
@login_required
def brands():
    conn = get_db()
    cur = conn.execute('''SELECT b.*, (SELECT COUNT(*) FROM items WHERE category = b.name) as item_count
                          FROM brands b ORDER BY b.name''')
    brands = cur.fetchall()
    conn.close()
    return render_template('brands.html', brands=brands)

@app.route('/brands/create', methods=['POST'])
@login_required
def create_brand():
    name = request.form.get('name', '').strip()
    emoji = request.form.get('emoji', '📦').strip()
    color = request.form.get('color', '⚪').strip()
    if not name:
        flash('Brand name required', 'danger')
        return redirect(url_for('brands'))
    conn = get_db()
    try:
        conn.execute("INSERT INTO brands (name, emoji, color) VALUES (?,?,?)",
                     (name, emoji, color))
        conn.commit()
        flash(f'Brand "{name}" created', 'success')
    except sqlite3.IntegrityError:
        flash('Brand already exists', 'danger')
    conn.close()
    return redirect(url_for('brands'))

@app.route('/brands/delete/<int:brand_id>', methods=['POST'])
@login_required
def delete_brand(brand_id):
    conn = get_db()
    conn.execute("DELETE FROM brands WHERE id = ?", (brand_id,))
    conn.commit()
    conn.close()
    flash('Brand deleted', 'success')
    return redirect(url_for('brands'))

@app.route('/groups')
@login_required
def groups():
    conn = get_db()
    cur = conn.execute('''SELECT pg.*, (SELECT COUNT(*) FROM items WHERE group_name = pg.name) as item_count
                          FROM product_groups pg ORDER BY pg.name''')
    groups = cur.fetchall()
    conn.close()
    return render_template('groups.html', groups=groups)

@app.route('/groups/create', methods=['POST'])
@login_required
def create_group():
    name = request.form.get('name', '').strip()
    emoji = request.form.get('emoji', '📦').strip()
    color = request.form.get('color', '⚪').strip()
    if not name:
        flash('Group name required', 'danger')
        return redirect(url_for('groups'))
    conn = get_db()
    try:
        conn.execute("INSERT INTO product_groups (name, emoji, color) VALUES (?,?,?)",
                     (name, emoji, color))
        conn.commit()
        flash(f'Group "{name}" created', 'success')
    except sqlite3.IntegrityError:
        flash('Group already exists', 'danger')
    conn.close()
    return redirect(url_for('groups'))

@app.route('/groups/delete/<int:group_id>', methods=['POST'])
@login_required
def delete_group(group_id):
    conn = get_db()
    conn.execute("DELETE FROM product_groups WHERE id = ?", (group_id,))
    conn.commit()
    conn.close()
    flash('Group deleted', 'success')
    return redirect(url_for('groups'))

@app.route('/categories')
@login_required
def categories():
    conn = get_db()
    cur = conn.execute('''SELECT pg.*, (SELECT COUNT(*) FROM items WHERE group_name = pg.name) as item_count
                          FROM product_groups pg ORDER BY pg.name''')
    categories = cur.fetchall()
    cur = conn.execute("SELECT DISTINCT category FROM items WHERE category IS NOT NULL AND category != '' ORDER BY category")
    brands = [r[0] for r in cur.fetchall()]
    conn.close()
    return render_template('categories.html', categories=categories, brands=brands)

@app.route('/categories/<int:group_id>')
@login_required
def category_products(group_id):
    conn = get_db()
    cur = conn.execute("SELECT * FROM product_groups WHERE id = ?", (group_id,))
    cat = cur.fetchone()
    if not cat:
        flash('Category not found', 'danger')
        return redirect(url_for('categories'))
    cur = conn.execute("SELECT i.*, b.emoji as brand_emoji FROM items i LEFT JOIN brands b ON i.category = b.name WHERE i.group_name = ? ORDER BY i.name", (cat['name'],))
    items = cur.fetchall()
    cur = conn.execute("SELECT DISTINCT category FROM items WHERE category IS NOT NULL AND category != '' ORDER BY category")
    brands = [r[0] for r in cur.fetchall()]
    conn.close()
    return render_template('category_items.html', category=cat, items=items, brands=brands)

@app.route('/categories/<int:group_id>/add', methods=['POST'])
@login_required
def category_add_item(group_id):
    conn = get_db()
    cur = conn.execute("SELECT * FROM product_groups WHERE id = ?", (group_id,))
    cat = cur.fetchone()
    if not cat:
        flash('Category not found', 'danger')
        return redirect(url_for('categories'))
    name = request.form.get('name', '').strip()
    brand = request.form.get('brand', '').strip()
    qty = request.form.get('qty', '0')
    price = request.form.get('price', '')
    cost = request.form.get('cost_price', '')
    ram = request.form.get('ram', '').strip()
    rom = request.form.get('rom', '').strip()
    code = request.form.get('product_code', '').strip()
    barcode_val = request.form.get('barcode', '').strip()
    if not name:
        flash('Name is required', 'danger')
        return redirect(url_for('category_products', group_id=group_id))
    if not brand:
        flash('Brand is required', 'danger')
        return redirect(url_for('category_products', group_id=group_id))
    try:
        qty = int(qty)
    except:
        qty = 0
    try:
        price = float(price.replace('$','').replace(',','')) if price else None
    except:
        price = None
    try:
        cost = float(cost.replace('$','').replace(',','')) if cost else None
    except:
        cost = None
    # Auto-generate barcode if not provided
    if not barcode_val:
        import random
        barcode_val = f'BR-{random.randint(10000, 99999)}-{random.randint(100, 999)}'
    # Auto-create brand if missing
    cur = conn.execute("SELECT id FROM brands WHERE name = ?", (brand,))
    if not cur.fetchone() and brand:
        conn.execute("INSERT INTO brands (name, emoji, color) VALUES (?, '📦', '🔵')", (brand,))
    conn.execute("INSERT INTO items (name, category, quantity, price, cost_price, ram, rom, group_name, product_code, barcode) VALUES (?,?,?,?,?,?,?,?,?,?)",
                 (name, brand, qty, price, cost, ram, rom, cat['name'], code, barcode_val))
    conn.commit()
    conn.close()
    flash(f'✅ Added {name} (Barcode: {barcode_val})', 'success')
    return redirect(url_for('category_products', group_id=group_id))

@app.route('/categories/<int:group_id>/import', methods=['POST'])
@login_required
def category_import(group_id):
    conn = get_db()
    cur = conn.execute("SELECT * FROM product_groups WHERE id = ?", (group_id,))
    cat = cur.fetchone()
    if not cat:
        flash('Category not found', 'danger')
        return redirect(url_for('categories'))
    file = request.files.get('file')
    if not file or not file.filename:
        flash('Please select a file', 'danger')
        return redirect(url_for('category_products', group_id=group_id))
    if not file.filename.endswith('.xlsx'):
        flash('Only .xlsx files are supported', 'danger')
        return redirect(url_for('category_products', group_id=group_id))
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            flash('Excel file is empty', 'danger')
            return redirect(url_for('category_products', group_id=group_id))
        headers = [str(c.value or '').strip().lower() if c.value else '' for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
        col_map = {'name': None, 'brand': None, 'quantity': None, 'price': None,
                   'cost': None, 'ram': None, 'rom': None, 'code': None, 'barcode': None}
        for h in headers:
            for key in col_map:
                if key in h:
                    col_map[key] = headers.index(h)
        if col_map['name'] is None:
            flash('Excel must have a "Name" column', 'danger')
            return redirect(url_for('category_products', group_id=group_id))
        conn = get_db()
        imported = 0
        errors = []
        import random
        for ri, row in enumerate(rows, 2):
            try:
                name = str(row[col_map['name']] or '').strip()
                if not name:
                    errors.append(f'Row {ri}: Name is empty'); continue
                brand = str(row[col_map['brand']] or 'General').strip() if col_map['brand'] is not None else 'General'
                qty = 0
                if col_map['quantity'] is not None:
                    try: qty = int(float(str(row[col_map['quantity']] or 0)))
                    except: pass
                price = None
                if col_map['price'] is not None:
                    try: price = float(str(row[col_map['price']] or '').replace('$',''))
                    except: pass
                cost = None
                if col_map['cost'] is not None:
                    try: cost = float(str(row[col_map['cost']] or '').replace('$',''))
                    except: pass
                ram = str(row[col_map['ram']] or '').strip() if col_map['ram'] is not None else ''
                rom = str(row[col_map['rom']] or '').strip() if col_map['rom'] is not None else ''
                code = str(row[col_map['code']] or '').strip() if col_map['code'] is not None else ''
                barcode_val = str(row[col_map['barcode']] or '').strip() if col_map['barcode'] is not None else ''
                if not barcode_val:
                    barcode_val = f'BR-{random.randint(10000, 99999)}-{random.randint(100, 999)}'
                # Auto-create brand if missing
                cur = conn.execute("SELECT id FROM brands WHERE name = ?", (brand,))
                if not cur.fetchone() and brand:
                    conn.execute("INSERT INTO brands (name, emoji, color) VALUES (?, '📦', '🔵')", (brand,))
                conn.execute(
                    "INSERT INTO items (name, category, quantity, price, cost_price, ram, rom, group_name, product_code, barcode) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (name, brand, qty, price, cost, ram, rom, cat['name'], code, barcode_val)
                )
                imported += 1
            except Exception as e:
                errors.append(f'Row {ri}: {str(e)}')
        conn.commit()
        conn.close()
        msg = f'✅ Imported {imported} items into "{cat["name"]}"'
        if errors:
            msg += f' with {len(errors)} errors'
            for e in errors[:5]:
                msg += f'\n- {e}'
            if len(errors) > 5:
                msg += f'\n... and {len(errors)-5} more errors'
        flash(msg, 'success' if not errors else 'warning')
    except Exception as e:
        flash(f'Error reading Excel: {str(e)}', 'danger')
    return redirect(url_for('category_products', group_id=group_id))

@app.route('/categories/<int:group_id>/delete/<int:item_id>', methods=['POST'])
@login_required
def category_delete_item(group_id, item_id):
    conn = get_db()
    cur = conn.execute("SELECT quantity, cost_price FROM items WHERE id = ?", (item_id,))
    item = cur.fetchone()
    if item:
        log_stock_movement(conn, item_id, 'deletion', -item['quantity'], item['quantity'], 0, item['cost_price'], f'deleted from category')
    conn.execute("DELETE FROM items WHERE id = ?", (item_id,))
    conn.commit()
    conn.close()
    flash('Item deleted', 'success')
    return redirect(url_for('category_products', group_id=group_id))

@app.route('/categories/create', methods=['POST'])
@login_required
def create_category():
    name = request.form.get('name', '').strip()
    emoji = request.form.get('emoji', '📦').strip()
    color = request.form.get('color', '⚪').strip()
    if not name:
        flash('Category name required', 'danger')
        return redirect(url_for('categories'))
    conn = get_db()
    try:
        conn.execute("INSERT INTO product_groups (name, emoji, color) VALUES (?,?,?)",
                     (name, emoji, color))
        conn.commit()
        flash(f'Category "{name}" created', 'success')
    except sqlite3.IntegrityError:
        flash('Category already exists', 'danger')
    conn.close()
    return redirect(url_for('categories'))

@app.route('/categories/delete/<int:group_id>', methods=['POST'])
@login_required
def delete_category(group_id):
    conn = get_db()
    cur = conn.execute("SELECT name FROM product_groups WHERE id = ?", (group_id,))
    cat = cur.fetchone()
    if cat:
        conn.execute("UPDATE items SET group_name = NULL WHERE group_name = ?", (cat['name'],))
        conn.execute("DELETE FROM product_groups WHERE id = ?", (group_id,))
        conn.commit()
        flash('Category deleted', 'success')
    conn.close()
    return redirect(url_for('categories'))

@app.route('/staff')
@login_required
def staff_list():
    conn = get_db()
    cur = conn.execute('''SELECT s.*, (SELECT COUNT(*) FROM staff) as total
                          FROM staff s ORDER BY s.role, s.name''')
    staff = cur.fetchall()
    conn.close()
    return render_template('staff.html', staff=staff)

@app.route('/staff/create', methods=['POST'])
@login_required
def staff_create():
    name = request.form.get('name', '').strip()
    role = request.form.get('role', 'staff')
    pin = request.form.get('pin', '').strip()
    if not name or not pin:
        flash('Name and PIN required', 'danger')
        return redirect(url_for('staff_list'))
    if not pin.isdigit() or len(pin) < 4:
        flash('PIN must be at least 4 digits', 'danger')
        return redirect(url_for('staff_list'))
    conn = get_db()
    try:
        conn.execute("INSERT INTO staff (name, role, pin) VALUES (?,?,?)", (name, role, pin))
        conn.commit()
        flash(f'Staff "{name}" created', 'success')
    except sqlite3.IntegrityError:
        flash('Failed to create staff', 'danger')
    conn.close()
    return redirect(url_for('staff_list'))

@app.route('/staff/delete/<int:staff_id>', methods=['POST'])
@login_required
def staff_delete(staff_id):
    conn = get_db()
    conn.execute("DELETE FROM staff WHERE id = ?", (staff_id,))
    conn.commit()
    conn.close()
    flash('Staff deleted', 'success')
    return redirect(url_for('staff_list'))

@app.route('/export/inventory')
@login_required
def export_inventory():
    conn = get_db()
    cur = conn.execute('''SELECT i.name, i.category, i.group_name, i.ram, i.rom,
                          i.quantity, i.price, i.cost_price
                          FROM items i ORDER BY i.name''')
    rows = cur.fetchall()
    conn.close()
    headers = ['Name','Brand','Group','RAM','ROM','Stock','Retail Price','Cost Price','Investment']
    data = [(r['name'], r['category'], r['group_name'] or '', r['ram'] or '', r['rom'] or '',
             r['quantity'], r['price'] or 0, r['cost_price'] or 0,
             (r['cost_price'] or 0) * r['quantity']) for r in rows]
    fp = export_excel('inventory.xlsx', headers, data)
    return send_file(fp, as_attachment=True, download_name='inventory.xlsx')

@app.route('/export/customers')
@login_required
def export_customers():
    conn = get_db()
    cur = conn.execute('''SELECT c.name, c.phone, c.phone2, c.phone3, c.location,
                          (SELECT COALESCE(SUM(amount),0) FROM payment_schedules
                           WHERE customer_id = c.id AND status IN ('pending','partial')) as total_debt,
                          (SELECT COUNT(*) FROM payment_schedules
                           WHERE customer_id = c.id) as total_bills
                          FROM customers c ORDER BY c.name''')
    rows = cur.fetchall()
    conn.close()
    headers = ['Name','Phone 1','Phone 2','Phone 3','Location','Total Debt','Total Bills']
    data = [(r['name'], r['phone'] or '', r['phone2'] or '', r['phone3'] or '',
             r['location'] or '', r['total_debt'], r['total_bills']) for r in rows]
    fp = export_excel('customers.xlsx', headers, data)
    return send_file(fp, as_attachment=True, download_name='customers.xlsx')

@app.route('/export/sales')
@login_required
def export_sales():
    days = request.args.get('days', '7')
    try: days = int(days)
    except: days = 7
    start = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    conn = get_db()
    cur = conn.execute('''SELECT s.id, i.name, i.category, s.quantity,
                          COALESCE(CAST(NULLIF(s.sellout_price,'') AS REAL), i.price) as price,
                          i.cost_price, s.timestamp, s.imei
                          FROM sales s JOIN items i ON s.item_id = i.id
                          WHERE DATE(s.timestamp) >= ? ORDER BY s.timestamp DESC''', (start,))
    rows = cur.fetchall()
    conn.close()
    headers = ['ID','Item','Brand','Qty','Unit Price','Unit Cost','Revenue','Profit','IMEI','Date']
    data = [(r['id'], r['name'], r['category'], r['quantity'], r['price'] or 0,
             r['cost_price'] or 0, (r['price'] or 0) * r['quantity'],
             ((r['price'] or 0) - (r['cost_price'] or 0)) * r['quantity'],
             r['imei'] or '', r['timestamp'][:10]) for r in rows]
    fp = export_excel(f'sales_last_{days}d.xlsx', headers, data)
    return send_file(fp, as_attachment=True, download_name=f'sales_{days}d.xlsx')

@app.route('/export/schedules')
@login_required
def export_schedules():
    status = request.args.get('status', 'all')
    conn = get_db()
    query = '''SELECT ps.id, c.name, c.phone, ps.due_date, ps.amount, ps.status, ps.notes
               FROM payment_schedules ps JOIN customers c ON ps.customer_id = c.id'''
    params = []
    if status != 'all':
        query += ' WHERE ps.status = ?'
        params.append(status)
    query += ' ORDER BY ps.due_date ASC'
    cur = conn.execute(query, params)
    rows = cur.fetchall()
    conn.close()
    headers = ['ID','Customer','Phone','Due Date','Amount','Status','Notes']
    data = [(r['id'], r['name'], r['phone'] or '', r['due_date'], r['amount'],
             r['status'], r['notes'] or '') for r in rows]
    fp = export_excel(f'schedules_{status}.xlsx', headers, data)
    return send_file(fp, as_attachment=True, download_name=f'schedules_{status}.xlsx')

@app.route('/export/brands')
@login_required
def export_brands():
    conn = get_db()
    cur = conn.execute('''SELECT b.name, b.emoji, b.color,
                          (SELECT COUNT(*) FROM items WHERE category = b.name) as item_count
                          FROM brands b ORDER BY b.name''')
    rows = cur.fetchall()
    conn.close()
    headers = ['Brand','Emoji','Color','Items']
    data = [(r['name'], r['emoji'], r['color'], r['item_count']) for r in rows]
    fp = export_excel('brands.xlsx', headers, data)
    return send_file(fp, as_attachment=True, download_name='brands.xlsx')

@app.route('/export/groups')
@login_required
def export_groups():
    conn = get_db()
    cur = conn.execute('''SELECT pg.name, pg.emoji, pg.color,
                          (SELECT COUNT(*) FROM items WHERE group_name = pg.name) as item_count
                          FROM product_groups pg ORDER BY pg.name''')
    rows = cur.fetchall()
    conn.close()
    headers = ['Group','Emoji','Color','Items']
    data = [(r['name'], r['emoji'], r['color'], r['item_count']) for r in rows]
    fp = export_excel('groups.xlsx', headers, data)
    return send_file(fp, as_attachment=True, download_name='groups.xlsx')

@app.route('/export/staff')
@login_required
def export_staff():
    conn = get_db()
    cur = conn.execute("SELECT name, role, pin FROM staff ORDER BY role, name")
    rows = cur.fetchall()
    conn.close()
    headers = ['Name','Role','PIN']
    data = [(r['name'], r['role'], r['pin']) for r in rows]
    fp = export_excel('staff.xlsx', headers, data)
    return send_file(fp, as_attachment=True, download_name='staff.xlsx')

@app.route('/exports')
@login_required
def exports_list():
    files = []
    if os.path.isdir(EXPORTS_DIR):
        for f in sorted(os.listdir(EXPORTS_DIR), reverse=True):
            fpath = os.path.join(EXPORTS_DIR, f)
            if os.path.isfile(fpath):
                size = os.path.getsize(fpath)
                files.append({'name': f, 'size': size, 'modified': datetime.fromtimestamp(os.path.getmtime(fpath)).strftime('%Y-%m-%d %H:%M')})
    return render_template('exports.html', files=files)

@app.route('/exports/download/<path:filename>')
@login_required
def export_download(filename):
    safe = os.path.basename(filename)
    fpath = os.path.join(EXPORTS_DIR, safe)
    if not os.path.isfile(fpath):
        abort(404)
    return send_file(fpath, as_attachment=True, download_name=safe)

@app.route('/inventory/download-template')
@login_required
def download_template():
    headers = ['Name', 'Brand', 'Quantity', 'Price', 'Cost', 'RAM', 'ROM', 'Group', 'Code', 'Barcode']
    rows = [['iPhone 14', 'Apple', 5, 999, 800, '6GB', '128GB', 'Phone', 'IP14-128', 'BR-12345-678'],
            ['Galaxy S23', 'Samsung', 3, 899, 700, '8GB', '256GB', 'Phone', 'S23-256', 'BR-54321-876']]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Import Template'
    hfont = Font(bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = Alignment(horizontal='center')
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    # Add data validation dropdown for Brand and Group
    conn = get_db()
    brands = [r['name'] for r in conn.execute("SELECT name FROM brands ORDER BY name").fetchall()]
    groups = [r['name'] for r in conn.execute("SELECT name FROM product_groups ORDER BY name").fetchall()]
    if brands:
        dv_brand = DataValidation(type="list", formula1='"'+','.join(brands)+'"', allow_blank=True)
        dv_brand.error = 'Please select a valid brand'
        dv_brand.errorTitle = 'Invalid Brand'
        ws.add_data_validation(dv_brand)
        dv_brand.add('B2:B1048576')
    if groups:
        dv_group = DataValidation(type="list", formula1='"'+','.join(groups)+'"', allow_blank=True)
        dv_group.error = 'Please select a valid group'
        dv_group.errorTitle = 'Invalid Group'
        ws.add_data_validation(dv_group)
        dv_group.add('I2:I1048576')
    # Auto-create brands/groups during import
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filepath = os.path.join(EXPORTS_DIR, f'{ts}_import_template.xlsx')
    wb.save(filepath)
    return send_file(filepath, as_attachment=True, download_name='import_template.xlsx')

@app.route('/inventory/import-excel', methods=['POST'])
@login_required
def import_excel():
    file = request.files.get('file')
    if not file or not file.filename:
        flash('Please select a file', 'danger')
        return redirect(url_for('inventory'))
    if not file.filename.endswith('.xlsx'):
        flash('Only .xlsx files are supported', 'danger')
        return redirect(url_for('inventory'))
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            flash('Excel file is empty', 'danger')
            return redirect(url_for('inventory'))
        headers = [str(c.value or '').strip().lower() if c.value else '' for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
        col_map = {'name': None, 'brand': None, 'quantity': None, 'price': None,
                   'cost': None, 'ram': None, 'rom': None, 'group': None, 'code': None, 'barcode': None}
        for h in headers:
            for key in col_map:
                if key in h:
                    col_map[key] = headers.index(h)
        if col_map['name'] is None:
            flash('Excel must have a "Name" column', 'danger')
            return redirect(url_for('inventory'))
        conn = get_db()
        imported = 0
        errors = []
        for ri, row in enumerate(rows, 2):
            try:
                name = str(row[col_map['name']] or '').strip()
                if not name:
                    errors.append(f'Row {ri}: Name is empty'); continue
                brand = str(row[col_map['brand']] or 'General').strip() if col_map['brand'] is not None else 'General'
                qty = 0
                if col_map['quantity'] is not None:
                    try: qty = int(float(str(row[col_map['quantity']] or 0)))
                    except: pass
                price = None
                if col_map['price'] is not None:
                    try: price = float(str(row[col_map['price']] or '').replace('$',''))
                    except: pass
                cost = None
                if col_map['cost'] is not None:
                    try: cost = float(str(row[col_map['cost']] or '').replace('$',''))
                    except: pass
                ram = str(row[col_map['ram']] or '').strip() if col_map['ram'] is not None else ''
                rom = str(row[col_map['rom']] or '').strip() if col_map['rom'] is not None else ''
                group = str(row[col_map['group']] or '').strip() if col_map['group'] is not None else ''
                code = str(row[col_map['code']] or '').strip() if col_map['code'] is not None else ''
                barcode_val = str(row[col_map['barcode']] or '').strip() if col_map['barcode'] is not None else ''
                if not barcode_val:
                    import random
                    barcode_val = f'BR-{random.randint(10000, 99999)}-{random.randint(100, 999)}'
                # Auto-create brand if missing
                cur = conn.execute("SELECT id FROM brands WHERE name = ?", (brand,))
                if not cur.fetchone() and brand:
                    conn.execute("INSERT INTO brands (name, emoji, color) VALUES (?, '📦', '🔵')", (brand,))
                # Auto-create group if missing
                if group:
                    cur = conn.execute("SELECT id FROM product_groups WHERE name = ?", (group,))
                    if not cur.fetchone():
                        conn.execute("INSERT INTO product_groups (name, emoji, color) VALUES (?, '📦', '🔵')", (group,))
                conn.execute(
                    "INSERT INTO items (name, category, quantity, price, cost_price, ram, rom, group_name, product_code, barcode) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (name, brand, qty, price, cost, ram, rom, group, code, barcode_val)
                )
                new_id = conn.lastrowid
                if new_id and qty > 0:
                    log_stock_movement(conn, new_id, 'import', qty, 0, qty, cost, 'Excel import', created_by=session.get('staff_name', ''))
                imported += 1
            except Exception as e:
                errors.append(f'Row {ri}: {str(e)}')
        conn.commit()
        conn.close()
        msg = f'✅ Imported {imported} items successfully'
        if errors:
            msg += f' with {len(errors)} errors'
            for e in errors[:5]:
                msg += f'\n- {e}'
            if len(errors) > 5:
                msg += f'\n... and {len(errors)-5} more errors'
        flash(msg, 'success' if not errors else 'warning')
        # Send Telegram import alert
        try:
            token = os.environ.get('TELEGRAM_BOT_TOKEN', '')
            if token:
                conn2 = get_db()
                target_ids = [r[0] for r in conn2.execute("SELECT chat_id FROM notification_users WHERE import_alerts=1 AND active=1").fetchall()]
                conn2.close()
                for cid in target_ids:
                    try:
                        import httpx
                        httpx.post(f'https://api.telegram.org/bot{token.split(",")[0].strip()}/sendMessage', json={
                            'chat_id': cid, 'parse_mode': 'Markdown',
                            'text': f'📦 *Import Alert*\n{imported} item(s) imported via webapp'
                        }, timeout=5)
                    except:
                        pass
        except:
            pass
    except Exception as e:
        flash(f'Error reading Excel file: {str(e)}', 'danger')
    return redirect(url_for('inventory'))

@app.route('/scan', methods=['GET', 'POST'])
@login_required
def scan():
    if request.method == 'POST':
        value = request.form.get('value', '').strip()
        if not value:
            flash('Scan a barcode or enter a code', 'warning')
            return redirect(url_for('scan'))
        conn = get_db()
        # Check IMEI first
        cur = conn.execute("SELECT item_id FROM item_imeis WHERE imei = ?", (value,))
        row = cur.fetchone()
        if row:
            conn.close()
            return redirect(url_for('item_detail', item_id=row['item_id']))
        # Check product code
        cur = conn.execute("SELECT id FROM items WHERE product_code = ?", (value,))
        row = cur.fetchone()
        if row:
            conn.close()
            return redirect(url_for('item_detail', item_id=row['id']))
        conn.close()
        flash(f'No item found for "{value}"', 'warning')
        return redirect(url_for('scan'))
    return render_template('scan.html')

@app.route('/inventory/barcode')
@login_required
def barcode_designer():
    conn = get_db()
    cur = conn.execute("SELECT i.*, b.emoji as brand_emoji, b.name as brand_name FROM items i LEFT JOIN brands b ON i.category = b.name ORDER BY b.name, i.name")
    items = cur.fetchall()
    cur = conn.execute("SELECT name FROM brands ORDER BY name")
    brands = cur.fetchall()
    conn.close()
    grouped = {}
    for item in items:
        brand = item['brand_name'] or item['category'] or 'Other'
        if brand not in grouped:
            grouped[brand] = []
        grouped[brand].append(item)
    layouts = [
        {'id': 'a4_5x9', 'name': 'A4 - 5×9 (38×28mm)', 'page_w': 210, 'page_h': 297, 'cols': 5, 'rows': 9, 'lw': 38, 'lh': 28, 'margin': 5, 'gap': 2},
        {'id': 'a4_4x13', 'name': 'A4 - 4×13 (50×20mm)', 'page_w': 210, 'page_h': 297, 'cols': 4, 'rows': 13, 'lw': 50, 'lh': 20, 'margin': 2, 'gap': 2},
        {'id': 'a4_3x10', 'name': 'A4 - 3×10 (63×27mm)', 'page_w': 210, 'page_h': 297, 'cols': 3, 'rows': 10, 'lw': 63, 'lh': 27, 'margin': 5, 'gap': 3},
        {'id': 'a4_2x7', 'name': 'A4 - 2×7 (100×38mm)', 'page_w': 210, 'page_h': 297, 'cols': 2, 'rows': 7, 'lw': 100, 'lh': 38, 'margin': 5, 'gap': 3},
        {'id': 'a5_2x7', 'name': 'A5 - 2×7 (74×38mm)', 'page_w': 148, 'page_h': 210, 'cols': 2, 'rows': 7, 'lw': 74, 'lh': 38, 'margin': 3, 'gap': 2},
        {'id': 'label_3x4', 'name': 'Label 3×4 (66×50mm)', 'page_w': 210, 'page_h': 297, 'cols': 3, 'rows': 4, 'lw': 66, 'lh': 50, 'margin': 5, 'gap': 3},
        {'id': 'label_2x5', 'name': 'Label 2×5 (100×50mm)', 'page_w': 210, 'page_h': 297, 'cols': 2, 'rows': 5, 'lw': 100, 'lh': 50, 'margin': 5, 'gap': 3},
    ]
    return render_template('barcode_designer.html', items=items, brands=brands, grouped=grouped, layouts=layouts)

@app.route('/inventory/barcode/<path:data>/svg')
@login_required
def barcode_svg(data):
    try:
        rv = barcode.Code128(data, writer=SVGWriter()).render()
        return rv, 200, {'Content-Type': 'image/svg+xml'}
    except Exception as e:
        return f'<svg width="200" height="80"><text x="10" y="40" font-size="12" fill="red">Error: {e}</text></svg>', 200, {'Content-Type': 'image/svg+xml'}

@app.route('/inventory/barcode/export-pdf', methods=['POST'])
@login_required
def barcode_export_pdf():
    import tempfile
    from fpdf import FPDF
    data = request.get_json(silent=True) or {}
    slots = data.get('slots', [])
    encode = data.get('encode', 'code')
    layout_id = data.get('layout', 'a4_5x9')
    layouts_map = {
        'a4_5x9': (210, 297, 5, 9, 38.0, 28.0, 5, 2),
        'a4_4x13': (210, 297, 4, 13, 50.0, 20.0, 2, 2),
        'a4_3x10': (210, 297, 3, 10, 63.0, 26.0, 5, 2),
        'a4_2x7': (210, 297, 2, 7, 100.0, 38.0, 5, 3),
        'a5_2x7': (148, 210, 2, 7, 72.0, 27.0, 3, 1),
        'label_3x4': (210, 297, 3, 4, 66.0, 50.0, 5, 3),
        'label_2x5': (210, 297, 2, 5, 100.0, 50.0, 5, 3),
    }
    pw, ph, COLS, ROWS, LABEL_W, LABEL_H, MARGIN, GAP = layouts_map.get(layout_id, (210, 297, 4, 13, 50.0, 20.0, 2, 2))
    is_landscape = pw > ph
    pdf = FPDF(orientation='L' if is_landscape else 'P', unit='mm', format=(pw, ph))
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()
    for idx, slot in enumerate(slots):
        col = idx % COLS
        row = (idx % (COLS * ROWS)) // COLS
        if idx > 0 and idx % (COLS * ROWS) == 0:
            pdf.add_page()
        x = MARGIN + col * (LABEL_W + GAP)
        y = MARGIN + row * (LABEL_H + GAP)
        pdf.set_draw_color(180)
        pdf.rect(x, y, LABEL_W, LABEL_H)
        content = slot.get('code', '') if encode == 'code' and slot.get('code') else slot.get('name', '')
        if content:
            try:
                svg = barcode.Code128(content, writer=SVGWriter()).render()
                tmp = tempfile.NamedTemporaryFile(suffix='.svg', delete=False)
                tmp.write(svg.encode('utf-8') if isinstance(svg, str) else svg)
                tmp.close()
                pdf.image(tmp.name, x=x+2, y=y+3, w=LABEL_W-4, h=9)
                os.unlink(tmp.name)
            except Exception:
                pass
            pdf.set_font('Helvetica', 'B', 7)
        pdf.set_xy(x+1, y+LABEL_H-5)
        name = f"{slot.get('brand','')} {slot.get('name','')}"
        try:
            pdf.cell(LABEL_W-2, 4, name[:30], align='C')
        except Exception:
            pdf.set_font('Helvetica', '', 6)
            pdf.set_xy(x+1, y+LABEL_H-5)
            safe = name.encode('latin-1', errors='replace').decode('latin-1')
            pdf.cell(LABEL_W-2, 4, safe[:30], align='C')
    pdf_bytes = pdf.output()
    return send_file(io.BytesIO(pdf_bytes), mimetype='application/pdf', as_attachment=True, download_name='barcodes.pdf')

def reverse_customer_debt(conn, sale):
    try:
        receipt_id = sale['receipt_id']
        if not receipt_id:
            return
        cur = conn.execute("SELECT customer_id FROM receipts WHERE id = ?", (receipt_id,))
        receipt = cur.fetchone()
        if not receipt or not receipt['customer_id']:
            return
        cust_id = receipt['customer_id']
        qty = sale['quantity'] or 1
        amount = qty * (float(sale['sellout_price'] or 0) + float(sale['delivery_fee'] or 0))
        if amount <= 0:
            return
        conn.execute("UPDATE receipts SET total_amount = MAX(0, total_amount - ?) WHERE id = ?", (amount, receipt_id))
        conn.execute("UPDATE payment_schedules SET amount = MAX(0, amount - ?) WHERE id = (SELECT id FROM payment_schedules WHERE customer_id = ? AND status = 'pending' ORDER BY due_date ASC LIMIT 1)", (amount, cust_id))
        conn.execute("UPDATE customers SET credit = MAX(0, credit - ?) WHERE id = ?", (amount, cust_id))
    except Exception:
        pass

@app.route('/sellout', methods=['GET'])
@login_required
def sellout_page():
    conn = get_db()
    cur = conn.execute("SELECT DISTINCT category FROM items WHERE quantity > 0 ORDER BY category")
    brands = [r['category'] for r in cur.fetchall()]
    cur = conn.execute("SELECT id, name, category, price FROM items WHERE quantity > 0 ORDER BY name")
    products = cur.fetchall()
    conn.close()
    return render_template('sellout.html', brands=brands, products=products)

@app.route('/customer-sellout/<int:cust_id>', methods=['GET'])
@login_required
def customer_sellout_page(cust_id):
    conn = get_db()
    cur = conn.execute("SELECT * FROM customers WHERE id = ?", (cust_id,))
    customer = cur.fetchone()
    conn.close()
    if not customer:
        flash('Customer not found', 'danger')
        return redirect(url_for('customers'))
    now_due = (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')
    return render_template('customer_sellout.html', customer=customer, now_due=now_due)

@app.route('/wholesale-sellout')
@login_required
def wholesale_sellout():
    conn = get_db()
    cur = conn.execute("SELECT id, name, phone, location FROM customers ORDER BY name")
    customers = cur.fetchall()
    cur = conn.execute("SELECT DISTINCT category FROM items WHERE quantity > 0 ORDER BY category")
    brands = [r['category'] for r in cur.fetchall()]
    cur = conn.execute("SELECT id, name, category, price FROM items WHERE quantity > 0 ORDER BY name")
    products = cur.fetchall()
    conn.close()
    now_due = (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')
    return render_template('wholesale_sellout.html', customers=customers, brands=brands, products=products, now_due=now_due)

@app.route('/customer-sellout/<int:cust_id>/complete', methods=['POST'])
@login_required
def customer_sellout_complete(cust_id):
    import json
    data = request.get_json(silent=True) or {}
    cart = data.get('cart', [])
    due_date = data.get('due_date', '')
    notes = data.get('notes', '')
    if not cart:
        return {'ok': False, 'msg': 'Cart is empty'}
    if not due_date:
        due_date = (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')
    conn = get_db()
    cur = conn.execute("SELECT name FROM customers WHERE id = ?", (cust_id,))
    customer = cur.fetchone()
    if not customer:
        conn.close()
        return {'ok': False, 'msg': 'Customer not found'}
    sold = 0
    total_amount = 0
    errors = []
    sale_ids = []
    for entry in cart:
        try:
            item_id = int(entry.get('id', 0))
            qty = int(entry.get('qty', 1))
            price = float(entry.get('price', 0))
            delivery = float(entry.get('delivery', 0))
            imei = entry.get('imei', '')
            cur = conn.execute("SELECT name, quantity FROM items WHERE id = ?", (item_id,))
            stock_row = cur.fetchone()
            if not stock_row or stock_row['quantity'] < qty:
                errors.append(f'Item #{item_id} ({"insufficient stock" if stock_row else "not found"}): {stock_row["quantity"] if stock_row else 0} left, {qty} requested')
                continue
            for _ in range(qty):
                cur = conn.execute(
                    "INSERT INTO sales (item_id, category, quantity, sellout_price, delivery_fee, special_note, imei) VALUES (?,?,?,?,?,?,?)",
                    (item_id, entry.get('brand', ''), 1, str(price), str(delivery), '', imei)
                )
                sale_ids.append(cur.lastrowid)
                conn.execute("UPDATE sales SET created_by = ? WHERE id = ?", (session.get('staff_name', ''), cur.lastrowid))
                conn.execute("UPDATE items SET quantity = quantity - 1 WHERE id = ? AND quantity > 0", (item_id,))
                if imei:
                    conn.execute("UPDATE item_imeis SET sold = 1 WHERE imei = ?", (imei,))
            log_stock_movement(conn, item_id, 'sale', -qty, reference=f'customer #{cust_id}', price=price, created_by=session.get('staff_name', ''))
            conn.execute(
                "INSERT INTO customer_sales (customer_id, item_id, quantity, price_paid, note, sale_date, created_by) VALUES (?,?,?,?,?,DATE('now'), ?)",
                (cust_id, item_id, qty, price * qty, notes or None, session.get('staff_name', ''))
            )
            sold += qty
            total_amount += (price + delivery) * qty
        except Exception as e:
            errors.append(str(e))
    if sold > 0:
        from datetime import datetime
        import random
        now = datetime.now()
        receipt_no = f'RCP-{now.strftime("%Y%m%d")}-{now.strftime("%H%M%S")}-{random.randint(100,999)}'
        cur = conn.execute(
            "INSERT INTO receipts (receipt_no, customer_id, customer_name, total_amount) VALUES (?,?,?,?)",
            (receipt_no, cust_id, customer['name'], total_amount))
        receipt_id = cur.lastrowid
        for sid in sale_ids:
            conn.execute("UPDATE sales SET receipt_id = ? WHERE id = ?", (receipt_id, sid))
        conn.execute(
            "INSERT INTO payment_schedules (customer_id, due_date, amount, status, notes) VALUES (?,?,?,'pending',?)",
            (cust_id, due_date, total_amount, notes or None)
        )
        conn.execute(
            "INSERT INTO transactions (customer_id, amount, description) VALUES (?,?,?)",
            (cust_id, total_amount, f'Sellout - {sold} item(s)')
        )
        conn.execute("UPDATE customers SET credit = credit + ? WHERE id = ?", (total_amount, cust_id))
    conn.commit()
    conn.close()
    return {'ok': True, 'msg': f'Sold {sold} item(s) to {customer["name"]}, bill due {due_date}', 'sold': sold, 'receipt_id': receipt_id if sold > 0 else 0, 'errors': errors}

@app.route('/sellout/scan', methods=['POST'])
@login_required
def sellout_scan():
    code = request.form.get('code', '').strip()
    if not code:
        return {'ok': False, 'msg': 'No code provided'}
    conn = get_db()
    # Check barcode first
    cur = conn.execute("SELECT i.id, i.name, i.category, i.price, i.cost_price, i.quantity, i.group_name, b.emoji as brand_emoji FROM items i LEFT JOIN brands b ON i.category = b.name WHERE i.barcode = ? AND i.quantity > 0", (code,))
    item = cur.fetchone()
    if item:
        conn.close()
        return {'ok': True, 'code': code, 'item': {'id': item['id'], 'name': item['name'], 'brand': item['category'], 'brand_emoji': item['brand_emoji'] or '', 'price': item['price'], 'cost': item['cost_price'], 'group': item['group_name'] or ''}}
    # Check IMEI next
    cur = conn.execute("SELECT item_id, sold FROM item_imeis WHERE imei = ?", (code,))
    imei_row = cur.fetchone()
    if imei_row:
        if imei_row['sold']:
            conn.close()
            return {'ok': False, 'msg': 'IMEI already sold'}
        cur = conn.execute("SELECT i.id, i.name, i.category, i.price, i.cost_price, i.quantity, i.group_name, b.emoji as brand_emoji FROM items i LEFT JOIN brands b ON i.category = b.name WHERE i.id = ?", (imei_row['item_id'],))
        item = cur.fetchone()
        if item and item['quantity'] > 0:
            conn.close()
            return {'ok': True, 'imei': code, 'item': {'id': item['id'], 'name': item['name'], 'brand': item['category'], 'brand_emoji': item['brand_emoji'] or '', 'price': item['price'], 'cost': item['cost_price'], 'group': item['group_name'] or ''}}
        conn.close()
        return {'ok': False, 'msg': 'Item out of stock'}
    # Check product code
    cur = conn.execute("SELECT i.id, i.name, i.category, i.price, i.cost_price, i.quantity, i.group_name, b.emoji as brand_emoji FROM items i LEFT JOIN brands b ON i.category = b.name WHERE i.product_code = ? AND i.quantity > 0", (code,))
    items = cur.fetchall()
    conn.close()
    if len(items) == 1:
        item = items[0]
        return {'ok': True, 'code': code, 'item': {'id': item['id'], 'name': item['name'], 'brand': item['category'], 'brand_emoji': item['brand_emoji'] or '', 'price': item['price'], 'cost': item['cost_price'], 'group': item['group_name'] or ''}}
    elif len(items) > 1:
        choices = [{'id': it['id'], 'name': it['name'], 'brand': it['category'], 'price': it['price']} for it in items]
        return {'ok': True, 'multiple': True, 'choices': choices, 'code': code}
    return {'ok': False, 'msg': 'No item found'}

@app.route('/sellout/complete', methods=['POST'])
@login_required
def sellout_complete():
    import json
    data = request.get_json(silent=True) or {}
    cart = data.get('cart', [])
    if not cart:
        return {'ok': False, 'msg': 'Cart is empty'}
    conn = get_db()
    sold = 0
    total_amt = 0
    errors = []
    sale_ids = []
    for entry in cart:
        try:
            item_id = int(entry.get('id', 0))
            qty = int(entry.get('qty', 1))
            price = float(entry.get('price', 0))
            delivery = float(entry.get('delivery', 0))
            note = entry.get('note', '')
            imei = entry.get('imei', '')
            cur = conn.execute("SELECT name, quantity FROM items WHERE id = ?", (item_id,))
            stock_row = cur.fetchone()
            if not stock_row or stock_row['quantity'] < qty:
                errors.append(f'Item #{item_id} ({"insufficient stock" if stock_row else "not found"}): {stock_row["quantity"] if stock_row else 0} left, {qty} requested')
                continue
            for _ in range(qty):
                cur = conn.execute(
                    "INSERT INTO sales (item_id, category, quantity, sellout_price, delivery_fee, special_note, imei) VALUES (?,?,?,?,?,?,?)",
                    (item_id, entry.get('brand', ''), 1, str(price), str(delivery), note, imei)
                )
                sale_ids.append(cur.lastrowid)
                conn.execute("UPDATE sales SET created_by = ? WHERE id = ?", (session.get('staff_name', ''), cur.lastrowid))
                conn.execute("UPDATE items SET quantity = quantity - 1 WHERE id = ? AND quantity > 0", (item_id,))
                if imei:
                    conn.execute("UPDATE item_imeis SET sold = 1 WHERE imei = ?", (imei,))
            log_stock_movement(conn, item_id, 'sale', -qty, reference='sellout', price=price, created_by=session.get('staff_name', ''))
            sold += qty
            total_amt += (price + delivery) * qty
        except Exception as e:
            errors.append(str(e))
    if sold > 0:
        from datetime import datetime
        import random
        now = datetime.now()
        receipt_no = f'RCP-{now.strftime("%Y%m%d")}-{now.strftime("%H%M%S")}-{random.randint(100,999)}'
        cur = conn.execute(
            "INSERT INTO receipts (receipt_no, total_amount) VALUES (?,?)",
            (receipt_no, total_amt))
        receipt_id = cur.lastrowid
        for sid in sale_ids:
            conn.execute("UPDATE sales SET receipt_id = ? WHERE id = ?", (receipt_id, sid))
    conn.commit()
    conn.close()
    return {'ok': True, 'msg': f'Sold {sold} item(s)', 'sold': sold, 'receipt_id': receipt_id if sold > 0 else 0, 'errors': errors}

@app.route('/receipt/<int:receipt_id>')
@login_required
def view_receipt(receipt_id):
    conn = get_db()
    cur = conn.execute("SELECT * FROM receipts WHERE id = ?", (receipt_id,))
    receipt = cur.fetchone()
    if not receipt:
        conn.close()
        flash('Receipt not found', 'danger')
        return redirect(url_for('sales'))
    cur = conn.execute('''SELECT s.*, i.name as item_name, i.category as brand,
        b.emoji as brand_emoji FROM sales s
        LEFT JOIN items i ON s.item_id = i.id
        LEFT JOIN brands b ON i.category = b.name
        WHERE s.receipt_id = ? ORDER BY s.id''', (receipt_id,))
    items = cur.fetchall()
    conn.close()
    return render_template('receipt.html', receipt=receipt, items=items)

@app.route('/receipt/<int:receipt_id>/print')
@login_required
def print_receipt(receipt_id):
    conn = get_db()
    cur = conn.execute("SELECT * FROM receipts WHERE id = ?", (receipt_id,))
    receipt = cur.fetchone()
    if not receipt:
        conn.close()
        flash('Receipt not found', 'danger')
        return redirect(url_for('sales'))
    cur = conn.execute('''SELECT s.*, i.name as item_name, i.category as brand,
        b.emoji as brand_emoji, i.price as retail_price FROM sales s
        LEFT JOIN items i ON s.item_id = i.id
        LEFT JOIN brands b ON i.category = b.name
        WHERE s.receipt_id = ? ORDER BY s.id''', (receipt_id,))
    items = cur.fetchall()
    conn.close()
    return render_template('receipt_print.html', receipt=receipt, items=items)

@app.route('/receipts')
@login_required
def receipts_list():
    conn = get_db()
    days = request.args.get('days', '30')
    try: days = int(days)
    except: days = 30
    start = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    cur = conn.execute('''SELECT r.*, (SELECT COUNT(*) FROM sales WHERE receipt_id = r.id) as item_count
        FROM receipts r WHERE DATE(r.created_at) >= ? ORDER BY r.created_at DESC''', (start,))
    receipts = cur.fetchall()
    conn.close()
    return render_template('receipts_list.html', receipts=receipts, days=days)

@app.route('/stock-movements')
@login_required
def stock_movements():
    conn = get_db()
    days = request.args.get('days', '7')
    change_type = request.args.get('type', '')
    item_search = request.args.get('search', '')
    try: days = int(days)
    except: days = 7
    query = "SELECT * FROM stock_movements WHERE 1=1"
    params = []
    if days > 0:
        start = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
        query += " AND DATE(created_at) >= ?"
        params.append(start)
    if change_type:
        query += " AND change_type = ?"
        params.append(change_type)
    if item_search:
        query += " AND (item_name LIKE ? OR item_category LIKE ?)"
        params.extend([f'%{item_search}%', f'%{item_search}%'])
    query += " ORDER BY created_at DESC LIMIT 500"
    cur = conn.execute(query, params)
    movements = cur.fetchall()
    cur = conn.execute("SELECT DISTINCT change_type FROM stock_movements ORDER BY change_type")
    types = [r['change_type'] for r in cur.fetchall()]
    conn.close()
    return render_template('stock_movements.html', movements=movements, types=types, days=days, change_type=change_type, search=item_search)

@app.route('/sales/<int:sale_id>/delete', methods=['POST'])
@login_required
def delete_sale(sale_id):
    conn = get_db()
    cur = conn.execute("SELECT * FROM sales WHERE id = ?", (sale_id,))
    sale = cur.fetchone()
    if not sale:
        conn.close()
        flash('Sale not found', 'danger')
        return redirect(url_for('sales'))
    item_id = sale['item_id']
    qty = sale['quantity'] or 1
    imei = sale['imei'] or ''
    reverse_customer_debt(conn, sale)
    conn.execute("DELETE FROM sales WHERE id = ?", (sale_id,))
    conn.execute("UPDATE items SET quantity = quantity + ? WHERE id = ?", (qty, item_id))
    log_stock_movement(conn, item_id, 'restore', qty, reference=f'sale #{sale_id} deleted')
    if imei:
        conn.execute("UPDATE item_imeis SET sold = 0 WHERE imei = ?", (imei,))
    conn.commit()
    conn.close()
    flash(f'Sale #{sale_id} deleted, stock restored', 'success')
    return redirect(url_for('sales'))

@app.route('/sales/<int:sale_id>/return', methods=['POST'])
@login_required
def return_sale(sale_id):
    reason = request.form.get('reason', '').strip()
    conn = get_db()
    cur = conn.execute("SELECT s.*, i.name as item_name FROM sales s JOIN items i ON s.item_id = i.id WHERE s.id = ?", (sale_id,))
    sale = cur.fetchone()
    if not sale:
        conn.close()
        flash('Sale not found', 'danger')
        return redirect(url_for('sales'))
    if sale['returned']:
        conn.close()
        flash('Sale already returned', 'warning')
        return redirect(url_for('sales'))
    qty = sale['quantity'] or 1
    reverse_customer_debt(conn, sale)
    conn.execute("UPDATE sales SET returned = 1, returned_at = datetime('now'), return_reason = ? WHERE id = ?", (reason or None, sale_id))
    conn.execute("UPDATE items SET quantity = quantity + ? WHERE id = ?", (qty, sale['item_id']))
    log_stock_movement(conn, sale['item_id'], 'return', qty, reference=f'sale #{sale_id} returned', price=sale['sellout_price'] or 0)
    if sale['imei']:
        conn.execute("UPDATE item_imeis SET sold = 0 WHERE imei = ?", (sale['imei'],))
    conn.commit()
    conn.close()
    flash(f'Sale #{sale_id} returned, {qty} item(s) restocked', 'success')
    return redirect(url_for('sales'))

@app.route('/sales/<int:sale_id>/edit', methods=['POST'])
@login_required
def edit_sale(sale_id):
    conn = get_db()
    cur = conn.execute("SELECT * FROM sales WHERE id = ?", (sale_id,))
    sale = cur.fetchone()
    if not sale:
        conn.close()
        flash('Sale not found', 'danger')
        return redirect(url_for('sales'))
    try:
        new_price = float(request.form.get('price', sale['sellout_price'] or 0))
        new_delivery = float(request.form.get('delivery', sale['delivery_fee'] or 0))
        new_note = request.form.get('note', sale['special_note'] or '')
        new_qty = int(request.form.get('quantity', sale['quantity'] or 1))
    except ValueError:
        flash('Invalid values', 'danger')
        conn.close()
        return redirect(url_for('sales'))
    old_qty = sale['quantity'] or 1
    diff = new_qty - old_qty
    conn.execute(
        "UPDATE sales SET sellout_price = ?, delivery_fee = ?, special_note = ?, quantity = ? WHERE id = ?",
        (str(new_price), str(new_delivery), new_note, new_qty, sale_id)
    )
    if diff > 0:
        conn.execute("UPDATE items SET quantity = quantity - ? WHERE id = ?", (diff, sale['item_id']))
        log_stock_movement(conn, sale['item_id'], 'adjustment', -diff, reference=f'sale #{sale_id} qty edited')
    elif diff < 0:
        conn.execute("UPDATE items SET quantity = quantity + ? WHERE id = ?", (abs(diff), sale['item_id']))
        log_stock_movement(conn, sale['item_id'], 'adjustment', abs(diff), reference=f'sale #{sale_id} qty edited')
    conn.commit()
    conn.close()
    flash(f'Sale #{sale_id} updated', 'success')
    return redirect(url_for('sales'))

@app.route('/returns')
@login_required
def returns():
    conn = get_db()
    days = request.args.get('days', '30')
    try: days = int(days)
    except: days = 30
    start = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    cur = conn.execute('''SELECT s.*, i.name as item_name, i.category as brand,
        COALESCE(CAST(NULLIF(s.sellout_price,'') AS REAL), i.price) as price, i.cost_price
        FROM sales s JOIN items i ON s.item_id = i.id
        WHERE s.returned = 1 AND DATE(s.returned_at) >= ?
        ORDER BY s.returned_at DESC''', (start,))
    returns = cur.fetchall()
    conn.close()
    return render_template('returns.html', returns=returns, days=days)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
